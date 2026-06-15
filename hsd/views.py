import datetime
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.template.loader import render_to_string
from hsd.models import HSDReceipt, HSDIssue, HSDDailyStock
from masters.models import Rig, Vendor, Equipment
from core.decorators import supervisor_required, admin_required


def _get_rigs():
    rigs = list(Rig.objects.values_list('rig_name', flat=True).order_by('rig_name'))
    return rigs if rigs else ['PPE-1', 'PPE-2', 'PPE-3', 'PPE-4', 'PPE-5']

def _get_user_rigs(request):
    all_rigs = _get_rigs()
    try:
        return request.user.profile.filter_rigs(all_rigs)
    except Exception:
        return all_rigs
def _apply_user_rig_filter(qs, request):
    """Always restrict queryset to user assigned rigs."""
    try:
        profile = request.user.profile
        if profile.role == 'admin':
            return qs
        assigned = profile.get_assigned_rigs()
        if assigned:
            return qs.filter(rig__in=assigned)
    except Exception:
        pass
    return qs



def _build_filter(request, method='GET'):
    data  = request.GET if method == 'GET' else request.POST
    rig_f = data.get('rig', '').strip()
    from_f= data.get('from', '').strip()
    to_f  = data.get('to', datetime.date.today().isoformat()).strip()
    month_f = data.get('month', '').strip()
    return rig_f, from_f, to_f, month_f


def _apply_date_filter(qs, from_f, to_f, month_f):
    if month_f and len(month_f) == 7:
        import calendar
        y, m = month_f.split('-')
        last = calendar.monthrange(int(y), int(m))[1]
        qs = qs.filter(date__gte=f'{month_f}-01', date__lte=f'{month_f}-{last:02d}')
    else:
        if from_f: qs = qs.filter(date__gte=from_f)
        if to_f:   qs = qs.filter(date__lte=to_f)
    return qs


@login_required
def hsd_dashboard(request):
    rig_f, from_f, to_f, month_f = _build_filter(request)

    receipts_qs = _apply_date_filter(HSDReceipt.objects.all(), from_f, to_f, month_f)
    issues_qs   = _apply_date_filter(HSDIssue.objects.all(),   from_f, to_f, month_f)

    receipts_qs = _apply_user_rig_filter(receipts_qs, request)
    issues_qs   = _apply_user_rig_filter(issues_qs,   request)
    if rig_f:
        receipts_qs = receipts_qs.filter(rig=rig_f)
        issues_qs   = issues_qs.filter(rig=rig_f)

    receipt_stats = receipts_qs.aggregate(
        total_received = Sum('quantity_ltrs'),
        total_amount   = Sum('invoice_amount'),
        count          = Count('id'),
    )
    issue_stats = issues_qs.aggregate(
        total_issued = Sum('quantity_ltrs'),
        count        = Count('id'),
    )

    # Per-rig summary
    rig_receipt = receipts_qs.values('rig').annotate(
        received=Sum('quantity_ltrs')).order_by('rig')
    rig_issue = issues_qs.values('rig').annotate(
        issued=Sum('quantity_ltrs')).order_by('rig')

    # Combine rig summaries
    rig_map = {}
    for r in rig_receipt: rig_map[r['rig']] = {'rig': r['rig'], 'received': r['received'], 'issued': 0}
    for r in rig_issue:
        if r['rig'] in rig_map:
            rig_map[r['rig']]['issued'] = r['issued']
        else:
            rig_map[r['rig']] = {'rig': r['rig'], 'received': 0, 'issued': r['issued']}
    for v in rig_map.values():
        v['balance'] = float(v['received'] or 0) - float(v['issued'] or 0)
    rig_summary = sorted(rig_map.values(), key=lambda x: x['rig'])

    # Issue by purpose
    issue_by_purpose = issues_qs.values('purpose').annotate(
        total=Sum('quantity_ltrs'), count=Count('id')
    ).order_by('-total')

    # Latest stock entries
    stock_entries = HSDDailyStock.objects.order_by('-date')[:10]

    return render(request, 'hsd/dashboard.html', {
        'page_title':       'HSD Dashboard',
        'rigs':             _get_rigs(),
        'filters':          {'rig': rig_f, 'from': from_f, 'to': to_f, 'month': month_f},
        'receipt_stats':    receipt_stats,
        'issue_stats':      issue_stats,
        'rig_summary':      rig_summary,
        'issue_by_purpose': issue_by_purpose,
        'stock_entries':    stock_entries,
    })


@login_required
@supervisor_required
def hsd_add_receipt(request):
    rigs    = _get_user_rigs(request)
    vendors = Vendor.objects.filter(status='Active').order_by('vendor_name')

    if request.method == 'POST':
        rig           = request.POST.get('rig', '').strip()
        date          = request.POST.get('date', '').strip()
        receipt_no    = request.POST.get('receipt_no', '').strip()
        supplier_id   = request.POST.get('supplier', '').strip()
        supplier_name = request.POST.get('supplier_name', '').strip()
        vehicle_no    = request.POST.get('vehicle_no', '').strip()
        qty           = request.POST.get('quantity_ltrs', '0') or '0'
        rate          = request.POST.get('rate_per_ltr', '').strip() or None
        invoice_amt   = request.POST.get('invoice_amount', '').strip() or None
        received_by   = request.POST.get('received_by', '').strip()
        remarks       = request.POST.get('remarks', '').strip()

        if not rig or not date:
            messages.error(request, 'Rig, date, and quantity are required.')
            return redirect('hsd_add_receipt')

        supplier_obj = None
        if supplier_id:
            try: supplier_obj = Vendor.objects.get(pk=supplier_id)
            except Vendor.DoesNotExist: pass

        # ── Duplicate check ──────────────────────────────────────
        if HSDReceipt.objects.filter(
            date=date, rig=rig,
            quantity_ltrs=float(qty)
        ).exists():
            messages.error(request,
                f'Duplicate: A receipt of {qty} L for {rig} on {date} already exists.')
            return redirect('hsd_add_receipt')
        # ─────────────────────────────────────────────────────────

        HSDReceipt.objects.create(
            date=date, rig=rig, receipt_no=receipt_no,
            supplier=supplier_obj, supplier_name=supplier_name,
            vehicle_no=vehicle_no, quantity_ltrs=float(qty),
            rate_per_ltr=float(rate) if rate else None,
            invoice_amount=float(invoice_amt) if invoice_amt else None,
            received_by=received_by, remarks=remarks,
            created_by=request.user,
        )
        # Recompute stock for this day
        _recompute_stock(rig, date, request.user)
        messages.success(request, f'Receipt of <strong>{qty} L</strong> recorded for {rig} on {date}.')
        return redirect('hsd_receipts')

    import datetime as dt
    yesterday = (dt.date.today() - dt.timedelta(days=1)).isoformat()
    return render(request, 'hsd/add_receipt.html', {
        'page_title': 'Add HSD Receipt',
        'rigs':       rigs,
        'vendors':    vendors,
        'today':      datetime.date.today().isoformat(),
        'default_date': yesterday,
    })


@login_required
@supervisor_required
def hsd_add_issue(request):
    rigs      = _get_user_rigs(request)
    from masters.models import EquipmentDeployment
    equipment = list(Equipment.objects.filter(
        status__in=['Available', 'Deployed']
    ).order_by('equipment_type', 'equipment_no'))

    # Build deployment map: equipment_id -> rig name
    active_deps = EquipmentDeployment.objects.filter(
        end_date__isnull=True
    ).values('equipment_id', 'deployed_to', 'deploy_type')
    dep_map = {d['equipment_id']: d['deployed_to'] for d in active_deps if d['deploy_type'] == 'Rig'}

    # Tag each equipment with its rig (or ALL if not deployed to a rig)
    for eq in equipment:
        if eq.id in dep_map:
            eq.rig_tag = dep_map[eq.id]
        else:
            # Also check equipment_no for rig pattern as fallback
            import re
            m = re.search(r'PPE-\d+', eq.equipment_no, re.IGNORECASE)
            eq.rig_tag = m.group(0).upper() if m else 'ALL'

    purposes  = HSDIssue.PURPOSE_CHOICES

    if request.method == 'POST':
        rig       = request.POST.get('rig', '').strip()
        date      = request.POST.get('date', '').strip()
        purpose   = request.POST.get('purpose', 'Drilling')
        issued_to = request.POST.get('issued_to', '').strip()
        eq_id     = request.POST.get('equipment', '').strip()
        qty       = request.POST.get('quantity_ltrs', '0') or '0'
        issued_by = request.POST.get('issued_by', '').strip()
        remarks   = request.POST.get('remarks', '').strip()

        if not rig or not date:
            messages.error(request, 'Rig, date, and quantity are required.')
            return redirect('hsd_add_issue')

        eq_obj = None
        if eq_id:
            try: eq_obj = Equipment.objects.get(pk=eq_id)
            except Equipment.DoesNotExist: pass

        meter_start_raw = request.POST.get('meter_start', '').strip()
        meter_end_raw   = request.POST.get('meter_end', '').strip()
        meter_start = float(meter_start_raw) if meter_start_raw else None
        meter_end   = float(meter_end_raw)   if meter_end_raw   else None
        meter_hours = round(meter_end - meter_start, 1) if meter_start and meter_end else None

        # ── Duplicate check ──────────────────────────────────────────
        dup_qs = HSDIssue.objects.filter(
            date=date, rig=rig, purpose=purpose,
        )
        if eq_obj:
            dup_qs = dup_qs.filter(equipment=eq_obj)
        elif issued_to:
            dup_qs = dup_qs.filter(issued_to=issued_to)

        if dup_qs.exists():
            messages.error(request,
                f'Duplicate entry: A {purpose} issue for '
                f'"{issued_to or (str(eq_obj) if eq_obj else rig)}" '
                f'on {date} already exists. Edit the existing entry instead.')
            return redirect('hsd_add_issue')
        # ─────────────────────────────────────────────────────────────

        HSDIssue.objects.create(
            date=date, rig=rig, purpose=purpose,
            issued_to=issued_to if issued_to else (eq_obj.equipment_no if eq_obj else ''),
            equipment=eq_obj, quantity_ltrs=float(qty),
            issued_by=issued_by, remarks=remarks,
            meter_start=meter_start, meter_end=meter_end, meter_hours=meter_hours,
            created_by=request.user,
        )
        _recompute_stock(rig, date, request.user)
        messages.success(request, f'Issue of <strong>{qty} L</strong> recorded for {issued_to or rig} on {date}.')
        return redirect('hsd_issues')

    import datetime as dt
    yesterday = (dt.date.today() - dt.timedelta(days=1)).isoformat()
    return render(request, 'hsd/add_issue.html', {
        'page_title': 'Add HSD Issue',
        'rigs':       rigs,
        'equipment':  equipment,
        'purposes':   purposes,
        'today':      datetime.date.today().isoformat(),
        'default_date': yesterday,
    })


@login_required
def hsd_api_stock_balance(request):
    from django.http import JsonResponse
    rig  = request.GET.get('rig', '').strip()
    date = request.GET.get('date', '').strip()
    if not rig:
        return JsonResponse({'balance': None, 'date': None, 'opening': None})
    import datetime as dt
    # Get latest stock entry for this rig up to selected date
    check_date = date or dt.date.today().isoformat()
    stock = HSDDailyStock.objects.filter(
        rig=rig, date__lte=check_date
    ).order_by('-date').first()
    if stock:
        return JsonResponse({
            'balance':  float(stock.closing_stock),
            'opening':  float(stock.opening_stock),
            'date':     str(stock.date),
            'receipts': float(stock.receipts),
            'consumption': float(stock.consumption),
        })
    return JsonResponse({'balance': None, 'date': None, 'opening': None})

@login_required
def hsd_api_update_opening(request):
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    rig     = request.POST.get('rig', '').strip()
    date    = request.POST.get('date', '').strip()
    opening = request.POST.get('opening', '').strip()
    if not rig or not date or not opening:
        return JsonResponse({'error': 'Missing fields'}, status=400)
    try:
        opening_val = float(opening)
        stock, _ = HSDDailyStock.objects.get_or_create(
            rig=rig, date=date,
            defaults={'created_by': request.user}
        )
        stock.opening_stock = opening_val
        stock.recompute()
        stock.save()
        return JsonResponse({
            'success': True,
            'new_balance': float(stock.closing_stock),
            'opening': float(stock.opening_stock)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def _recompute_stock(rig, date, user):
    """Recompute stock for date AND cascade forward to all subsequent days."""
    import datetime as dt

    # Step 1 — Recompute the target date
    stock, _ = HSDDailyStock.objects.get_or_create(
        date=date, rig=rig,
        defaults={'created_by': user}
    )
    # Carry forward opening from previous day
    prev_date = (dt.date.fromisoformat(str(date)) - dt.timedelta(days=1)).isoformat()
    prev = HSDDailyStock.objects.filter(date=prev_date, rig=rig).first()
    if prev:
        stock.opening_stock = float(prev.closing_stock)
    stock.recompute()  # saves internally

    # Step 2 — Cascade forward: update all subsequent days for this rig
    future_stocks = HSDDailyStock.objects.filter(
        rig=rig, date__gt=date
    ).order_by('date')

    prev_closing = float(stock.closing_stock)
    for fs in future_stocks:
        fs.opening_stock = prev_closing
        fs.recompute()  # saves internally
        prev_closing = float(fs.closing_stock)


@login_required
@supervisor_required
def hsd_edit_stock(request, pk):
    from django.http import JsonResponse
    stock = get_object_or_404(HSDDailyStock, pk=pk)
    if request.method == 'POST':
        try:
            opening  = request.POST.get('opening_stock', '').strip()
            remarks  = request.POST.get('remarks', '').strip()
            if opening != '':
                stock.opening_stock = float(opening)
            stock.remarks = remarks
            stock.recompute()
            stock.save()
            messages.success(request, f'Stock updated for {stock.rig} on {stock.date}.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('hsd_stock')

@login_required
@admin_required
def hsd_delete_stock(request, pk):
    stock = get_object_or_404(HSDDailyStock, pk=pk)
    if request.method == 'POST':
        info = f'{stock.rig} on {stock.date}'
        stock.delete()
        messages.success(request, f'Stock entry for {info} deleted.')
    return redirect('hsd_stock')

@login_required
def hsd_receipts(request):
    rig_f, from_f, to_f, month_f = _build_filter(request)
    qs = _apply_date_filter(HSDReceipt.objects.all(), from_f, to_f, month_f)
    qs = _apply_user_rig_filter(qs, request)
    if rig_f: qs = qs.filter(rig=rig_f)
    qs = qs.select_related('supplier', 'created_by').order_by('-date', 'rig')

    stats = qs.aggregate(
        total_qty    = Sum('quantity_ltrs'),
        total_amount = Sum('invoice_amount'),
        count        = Count('id'),
    )
    return render(request, 'hsd/receipts.html', {
        'page_title': 'HSD Receipts',
        'receipts':   qs,
        'stats':      stats,
        'rigs':       _get_user_rigs(request),
        'filters':    {'rig': rig_f, 'from': from_f, 'to': to_f, 'month': month_f},
    })


@login_required
def hsd_issues(request):
    rig_f, from_f, to_f, month_f = _build_filter(request)
    qs = _apply_date_filter(HSDIssue.objects.all(), from_f, to_f, month_f)
    qs = _apply_user_rig_filter(qs, request)
    if rig_f: qs = qs.filter(rig=rig_f)
    qs = qs.select_related('equipment', 'created_by').order_by('-date', 'rig')

    stats = qs.aggregate(total_qty=Sum('quantity_ltrs'), count=Count('id'))
    by_purpose = qs.values('purpose').annotate(qty=Sum('quantity_ltrs')).order_by('-qty')

    return render(request, 'hsd/issues.html', {
        'page_title': 'HSD Issues',
        'issues':     qs,
        'stats':      stats,
        'by_purpose': by_purpose,
        'rigs':       _get_user_rigs(request),
        'filters':    {'rig': rig_f, 'from': from_f, 'to': to_f, 'month': month_f},
    })


@login_required
def hsd_stock(request):
    rig_f, from_f, to_f, month_f = _build_filter(request)
    qs = _apply_date_filter(HSDDailyStock.objects.all(), from_f, to_f, month_f)
    qs = _apply_user_rig_filter(qs, request)
    if rig_f: qs = qs.filter(rig=rig_f)
    qs = qs.order_by('-date', 'rig')

    from django.db.models import Sum, Avg
    import json

    stats = qs.aggregate(
        total_received = Sum('receipts'),
        total_issued   = Sum('consumption'),
        avg_closing    = Avg('closing_stock'),
    )

    COLORS = ['#3b82f6','#ef4444','#f59e0b','#22c55e','#8b5cf6','#14b8a6']
    chart_rigs = list(qs.values_list('rig', flat=True).distinct().order_by('rig'))
    all_dates  = list(qs.values_list('date', flat=True).distinct().order_by('date'))
    date_labels = json.dumps([str(d) for d in all_dates])

    trend_datasets = []
    for i, rig in enumerate(chart_rigs):
        rig_qs   = qs.filter(rig=rig).order_by('date')
        date_map = {str(s.date): float(s.closing_stock) for s in rig_qs}
        data     = [date_map.get(str(d), None) for d in all_dates]
        trend_datasets.append({'label': rig, 'data': data, 'color': COLORS[i % len(COLORS)]})

    rig_summary_raw = list(qs.values('rig').annotate(
        total_rcv=Sum('receipts'),
        total_iss=Sum('consumption'),
    ).order_by('rig'))

    # Calculate net balance for each rig
    rig_summary = []
    for r in rig_summary_raw:
        r['total_rcv'] = float(r['total_rcv'] or 0)
        r['total_iss'] = float(r['total_iss'] or 0)
        r['net_balance'] = round(r['total_rcv'] - r['total_iss'], 1)
        rig_summary.append(r)

    return render(request, 'hsd/stock.html', {
        'page_title':     'HSD Daily Stock',
        'stock':          qs,
        'rigs':           _get_rigs(),
        'filters':        {'rig': rig_f, 'from': from_f, 'to': to_f, 'month': month_f},
        'stats':          stats,
        'date_labels':    date_labels,
        'trend_datasets': json.dumps(trend_datasets),
        'rig_summary':    rig_summary,
        'chart_colors':   json.dumps(COLORS),
    })



@login_required
@admin_required
def hsd_delete_receipt(request, pk):
    obj = get_object_or_404(HSDReceipt, pk=pk)
    if request.method == 'POST':
        rig, date = obj.rig, obj.date
        obj.delete()
        _recompute_stock(rig, str(date), request.user)
        messages.success(request, 'Receipt deleted.')
    return redirect('hsd_receipts')


@login_required
@admin_required
def hsd_delete_issue(request, pk):
    obj = get_object_or_404(HSDIssue, pk=pk)
    if request.method == 'POST':
        rig, date = obj.rig, obj.date
        obj.delete()
        _recompute_stock(rig, str(date), request.user)
        messages.success(request, 'Issue deleted.')
    return redirect('hsd_issues')


@login_required
def hsd_export_excel(request):
    if request.method != 'POST':
        return redirect('hsd_dashboard')

    rig_f, from_f, to_f, month_f = _build_filter(request, 'POST')
    receipts = _apply_date_filter(HSDReceipt.objects.all(), from_f, to_f, month_f)
    issues   = _apply_date_filter(HSDIssue.objects.all(),   from_f, to_f, month_f)
    stock    = _apply_date_filter(HSDDailyStock.objects.all(), from_f, to_f, month_f)
    if rig_f:
        receipts = receipts.filter(rig=rig_f)
        issues   = issues.filter(rig=rig_f)
        stock    = stock.filter(rig=rig_f)

    wb  = openpyxl.Workbook()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    NAVY = PatternFill('solid', fgColor='0B3D6D')
    hf   = Font(bold=True, color='FFFFFF', size=10)

    def make_sheet(title, headers, rows_fn):
        ws = wb.create_sheet(title)
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = hf; c.fill = NAVY
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.row_dimensions[1].height = 25
        for idx, row in enumerate(rows_fn(), 2):
            bg = PatternFill('solid', fgColor='FFFFFF' if idx%2==0 else 'F8FAFC')
            for col, v in enumerate(row, 1):
                cell = ws.cell(idx, col, v)
                cell.fill = bg
                if isinstance(v, datetime.date):
                    cell.number_format = 'DD-MMM-YYYY'

    make_sheet('Receipts',
        ['Date','Rig','Receipt No','Supplier','Vehicle No','Qty (L)','Rate/L','Amount','Remarks'],
        lambda: [(r.date,r.rig,r.receipt_no,r.supplier_name or str(r.supplier or ''),
                  r.vehicle_no,float(r.quantity_ltrs),
                  float(r.rate_per_ltr) if r.rate_per_ltr else '',
                  float(r.invoice_amount) if r.invoice_amount else '',
                  r.remarks) for r in receipts.order_by('date','rig')]
    )
    make_sheet('Issues',
        ['Date','Rig','Purpose','Issued To','Qty (L)','Issued By','Remarks'],
        lambda: [(i.date,i.rig,i.purpose,i.issued_to,float(i.quantity_ltrs),
                  i.issued_by,i.remarks) for i in issues.order_by('date','rig')]
    )
    make_sheet('Stock',
        ['Date','Rig','Opening (L)','Received (L)','Issued (L)','Closing (L)'],
        lambda: [(s.date,s.rig,float(s.opening_stock),float(s.receipts),
                  float(s.consumption),float(s.closing_stock))
                 for s in stock.order_by('date','rig')]
    )

    # -- SHEET 4: Daily Consumption by Purpose (time series + line chart) --
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import Border, Side
    from openpyxl.drawing.line import LineProperties

    PURPOSE_COLORS = {
        'Drilling':  '3B82F6',
        'Generator': 'EF4444',
        'Equipment': 'F59E0B',
        'Vehicle':   '22C55E',
        'Camp':      '8B5CF6',
        'Other':     '64748B',
    }

    daily_raw = list(
        issues.values('date', 'purpose')
              .annotate(qty=Sum('quantity_ltrs'))
              .order_by('date', 'purpose')
    )

    # Build map and collect active purposes/dates
    date_map = {}
    active_purposes = []
    for row in daily_raw:
        d = row['date']
        p = row['purpose']
        if d not in date_map:
            date_map[d] = {}
        date_map[d][p] = float(row['qty'] or 0)
        if p not in active_purposes:
            active_purposes.append(p)
    active_purposes = sorted(active_purposes)
    sorted_dates    = sorted(date_map.keys())

    if sorted_dates and active_purposes:
        thin_s = Side(style='thin', color='E2E8F0')
        bdr_s  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

        ws4 = wb.create_sheet('Daily Consumption Chart')

        # -- Header --
        ws4.cell(1, 1, 'Date').font      = Font(bold=True, color='FFFFFF', size=10)
        ws4.cell(1, 1).fill              = PatternFill('solid', fgColor='0B3D6D')
        ws4.cell(1, 1).alignment         = Alignment(horizontal='center')
        ws4.cell(1, 1).border            = bdr_s
        ws4.column_dimensions['A'].width = 14
        ws4.row_dimensions[1].height     = 25

        for ci, purpose in enumerate(active_purposes, 2):
            c = ws4.cell(1, ci, purpose)
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', fgColor='0B3D6D')
            c.alignment = Alignment(horizontal='center')
            c.border    = bdr_s
            ws4.column_dimensions[get_column_letter(ci)].width = 13

        ws4.freeze_panes = 'B2'

        # -- Data rows --
        for ri, d in enumerate(sorted_dates, 2):
            even = ri % 2 == 0
            bg   = PatternFill('solid', fgColor='FFFFFF' if even else 'F0F9FF')

            dc = ws4.cell(ri, 1, d)
            dc.number_format = 'DD-MMM-YY'
            dc.font          = Font(bold=True, size=9, color='1E293B')
            dc.fill          = bg
            dc.border        = bdr_s
            dc.alignment     = Alignment(horizontal='center')

            for ci, purpose in enumerate(active_purposes, 2):
                qty = date_map[d].get(purpose, None)
                c   = ws4.cell(ri, ci, qty)
                c.fill      = bg
                c.border    = bdr_s
                c.alignment = Alignment(horizontal='center')
                if qty:
                    c.font = Font(bold=True, size=9,
                                  color=PURPOSE_COLORS.get(purpose, '1E293B'))
                    c.number_format = '0.0'

        # -- Totals row --
        total_row = len(sorted_dates) + 2
        tc        = ws4.cell(total_row, 1, 'TOTAL')
        tc.font   = Font(bold=True, color='FFFFFF', size=10)
        tc.fill   = PatternFill('solid', fgColor='0B3D6D')
        tc.border = bdr_s
        tc.alignment = Alignment(horizontal='center')

        for ci, purpose in enumerate(active_purposes, 2):
            cl = get_column_letter(ci)
            c  = ws4.cell(total_row, ci,
                          f'=SUM({cl}2:{cl}{total_row-1})')
            c.font        = Font(bold=True, color='FFFFFF', size=10)
            c.fill        = PatternFill('solid', fgColor='0B3D6D')
            c.border      = bdr_s
            c.alignment   = Alignment(horizontal='center')
            c.number_format = '0.0'

        # -- Line Chart --
        chart              = LineChart()
        chart.title        = 'Daily HSD Consumption by Purpose'
        chart.style        = 2
        chart.y_axis.title = 'Quantity (Litres)'
        chart.x_axis.title = 'Date'
        chart.height       = 16
        chart.width        = 32
        chart.grouping     = 'standard'
        chart.y_axis.numFmt = '0'
        chart.x_axis.numFmt = 'DD-MMM'

        # Categories = dates column (A2:A{last_data_row})
        last_data_row = total_row - 1
        cats = Reference(ws4, min_col=1, min_row=2, max_row=last_data_row)
        chart.set_categories(cats)

        # One series per purpose - values include header row so title is picked up
        for ci, purpose in enumerate(active_purposes, 2):
            vals   = Reference(ws4, min_col=ci, min_row=1, max_row=last_data_row)
            chart.add_data(vals, titles_from_data=True)

            s = chart.series[-1]
            color = PURPOSE_COLORS.get(purpose, '64748B')

            # Line style
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width     = 25000   # 2.5pt

            # Smooth curve
            s.smooth = True

            # Marker
            s.marker.symbol = 'circle'
            s.marker.size   = 5
            s.marker.graphicalProperties.solidFill             = color
            s.marker.graphicalProperties.line.solidFill        = color

        # Place chart starting 3 rows below totals
        ws4.add_chart(chart, f'A{total_row + 3}')

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="HSD_Report_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def hsd_export_pdf(request):
    if request.method != 'POST':
        return redirect('hsd_dashboard')

    rig_f, from_f, to_f, month_f = _build_filter(request, method='POST')
    receipts_qs = _apply_date_filter(HSDReceipt.objects.all(), from_f, to_f, month_f)
    issues_qs   = _apply_date_filter(HSDIssue.objects.all(),   from_f, to_f, month_f)
    receipts_qs = _apply_user_rig_filter(receipts_qs, request)
    issues_qs   = _apply_user_rig_filter(issues_qs,   request)
    if rig_f:
        receipts_qs = receipts_qs.filter(rig=rig_f)
        issues_qs   = issues_qs.filter(rig=rig_f)

    # Build rig-wise sections
    all_rigs = sorted(set(
        list(receipts_qs.values_list('rig', flat=True).distinct()) +
        list(issues_qs.values_list('rig', flat=True).distinct())
    ))

    rig_sections = []
    for rig in all_rigs:
        r_qs = receipts_qs.filter(rig=rig).order_by('date')
        i_qs = issues_qs.filter(rig=rig).select_related('equipment').order_by('date', 'purpose')

        total_received = float(r_qs.aggregate(t=Sum('quantity_ltrs'))['t'] or 0)
        total_issued   = float(i_qs.aggregate(t=Sum('quantity_ltrs'))['t'] or 0)

        # By purpose with percentage
        by_purpose_raw = list(
            i_qs.values('purpose')
                .annotate(qty=Sum('quantity_ltrs'))
                .order_by('-qty')
        )
        for p in by_purpose_raw:
            p['qty'] = float(p['qty'] or 0)
            p['pct'] = round(p['qty'] / total_issued * 100, 1) if total_issued > 0 else 0

        # Clean up issued_to - remove duplicate bracket part like "NAME (NAME)"
        def clean_issued_to(issues_list):
            import re
            result = []
            for i in issues_list:
                name = i.issued_to or ''
                # Remove pattern: "SOME_NAME (SOME_NAME)" or "SOME NAME (SOME_NAME)"
                # Strip registration numbers in brackets for cleaner display
                match = re.match(r'^([^(]+)\s*\([^)]+\)\s*$', name.strip())
                if match:
                    base = match.group(1).strip()
                    # Only remove bracket if bracket content is same/similar to base
                    bracket = re.search(r'\(([^)]+)\)', name)
                    if bracket:
                        bracket_text = bracket.group(1).strip()
                        # If bracket is registration number (has letters+digits pattern) keep it
                        if re.match(r'^[A-Z]{2}[0-9]{2}', bracket_text):
                            i.issued_to = name  # keep reg number
                        else:
                            i.issued_to = base  # remove duplicate name
                    else:
                        i.issued_to = base
                else:
                    i.issued_to = name
                result.append(i)
            return result

        rig_sections.append({
            'rig':            rig,
            'receipts':       r_qs,
            'issues':         clean_issued_to(list(i_qs)),
            'total_received': total_received,
            'total_issued':   total_issued,
            'balance':        total_received - total_issued,
            'receipt_count':  r_qs.count(),
            'issue_count':    i_qs.count(),
            'by_purpose':     by_purpose_raw,
        })

    import datetime as dt
    # Debug: log what data we have
    import logging
    logger = logging.getLogger(__name__)
    logger.warning(f'PDF: rig_f={rig_f} from_f={from_f} to_f={to_f} rigs={all_rigs} sections={len(rig_sections)}')

    html = render_to_string('exports/hsd_pdf.html', {
        'rig_sections': rig_sections,
        'filters':      {'rig': rig_f, 'from': from_f, 'to': to_f, 'month': month_f},
        'generated':    dt.datetime.now(),
    }, request=request)

    try:
        import weasyprint
        pdf = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback
        return HttpResponse(f'<pre>PDF Error:\n{traceback.format_exc()}</pre>', status=500)

    response = HttpResponse(pdf, content_type='application/pdf')
    fname = f"HSD_Diesel_{dt.date.today()}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
@supervisor_required
def hsd_edit_issue(request, pk):
    obj            = get_object_or_404(HSDIssue, pk=pk)
    rigs           = _get_rigs()
    equipment_list = Equipment.objects.filter(
        status__in=['Available', 'Deployed']
    ).order_by('equipment_type', 'equipment_no')
    purposes       = HSDIssue.PURPOSE_CHOICES

    if request.method == 'POST':
        date_val      = request.POST.get('date', '').strip()
        rig           = request.POST.get('rig', obj.rig)
        purpose       = request.POST.get('purpose', obj.purpose)
        issued_to     = request.POST.get('issued_to', '').strip()
        qty           = request.POST.get('quantity_ltrs', str(obj.quantity_ltrs))
        issued_by     = request.POST.get('issued_by', '').strip()
        remarks       = request.POST.get('remarks', '').strip()
        eq_id         = request.POST.get('equipment', '').strip()
        eq_obj        = Equipment.objects.filter(pk=eq_id).first() if eq_id else None

        # Duplicate check — exclude current record
        dup_qs = HSDIssue.objects.filter(
            date=date_val or obj.date,
            rig=rig,
            purpose=purpose,
            quantity_ltrs=float(qty),
        ).exclude(pk=pk)
        if eq_obj:
            dup_qs = dup_qs.filter(equipment=eq_obj)
        else:
            dup_qs = dup_qs.filter(issued_to=issued_to)

        if dup_qs.exists():
            messages.error(request,
                f'Duplicate entry: A {purpose} issue of {qty} L '
                f'for {issued_to or str(eq_obj)} on {date_val} already exists.')
            return redirect('hsd_issues')

        if date_val:
            obj.date      = date_val
        obj.rig           = rig
        obj.purpose       = purpose
        obj.issued_to     = issued_to or (str(eq_obj) if eq_obj else obj.issued_to)
        obj.quantity_ltrs = float(qty)
        obj.issued_by     = issued_by
        obj.remarks       = remarks
        obj.equipment     = eq_obj
        ms = request.POST.get('meter_start', '').strip()
        me = request.POST.get('meter_end', '').strip()
        obj.meter_start = float(ms) if ms else None
        obj.meter_end   = float(me) if me else None
        obj.meter_hours = round(float(me) - float(ms), 1) if ms and me else None
        obj.save()
        _recompute_stock(obj.rig, str(obj.date), request.user)
        messages.success(request, 'Issue updated successfully.')
        return redirect('hsd_issues')

    return render(request, 'hsd/edit_issue.html', {
        'page_title':     'Edit HSD Issue',
        'obj':            obj,
        'rigs':           rigs,
        'equipment_list': equipment_list,
        'purposes':       purposes,
    })


@login_required
def hsd_get_last_meter(request):
    """API: return last meter_end reading for an equipment."""
    from django.http import JsonResponse
    eq_id = request.GET.get('equipment_id', '').strip()
    if not eq_id:
        return JsonResponse({'meter_end': None})
    last = HSDIssue.objects.filter(
        equipment_id=eq_id,
        meter_end__isnull=False
    ).order_by('-date', '-id').first()
    return JsonResponse({'meter_end': float(last.meter_end) if last else None})



@login_required
@supervisor_required
def hsd_bulk_issue(request):
    """Bulk daily HSD issue — show all rig equipment in one table."""
    from masters.models import EquipmentDeployment
    import datetime as dt
    rigs      = _get_user_rigs(request)
    today     = dt.date.today().isoformat()
    yesterday = (dt.date.today() - dt.timedelta(days=1)).isoformat()

    if request.method == 'POST':
        rig      = request.POST.get('rig', '').strip()
        date     = request.POST.get('date', '').strip()
        remarks  = request.POST.get('remarks', '').strip()

        if not rig or not date:
            messages.error(request, 'Rig and date are required.')
            return redirect('hsd_bulk_issue')

        # Get all equipment rows submitted
        eq_ids    = request.POST.getlist('eq_id')
        eq_names  = request.POST.getlist('eq_name')
        eq_types  = request.POST.getlist('eq_type')
        quantities = request.POST.getlist('quantity')
        meter_starts = request.POST.getlist('meter_start')
        meter_ends   = request.POST.getlist('meter_end')
        purposes_list = request.POST.getlist('purpose')

        saved = 0
        for i in range(len(eq_names)):
            qty_raw = quantities[i].strip() if i < len(quantities) else '0'
            if not qty_raw:
                qty_raw = '0'
            try:
                qty = float(qty_raw)
            except ValueError:
                qty = 0.0

            ms_raw = meter_starts[i].strip() if i < len(meter_starts) else ''
            me_raw = meter_ends[i].strip()   if i < len(meter_ends)   else ''
            ms = float(ms_raw) if ms_raw else None
            me = float(me_raw) if me_raw else None
            mh = round(me - ms, 1) if ms and me else None

            eq_id  = eq_ids[i].strip() if i < len(eq_ids) else ''
            eq_name = eq_names[i].strip() if i < len(eq_names) else ''
            purpose = purposes_list[i].strip() if i < len(purposes_list) else 'Drilling'

            eq_obj = None
            if eq_id:
                try:
                    from masters.models import Equipment as Eq
                    eq_obj = Eq.objects.get(pk=int(eq_id))
                except Exception:
                    pass

            # Only save if qty > 0 OR meter readings entered
            if qty > 0 or (ms is not None and me is not None):
                # Check duplicate
                dup = HSDIssue.objects.filter(
                    date=date, rig=rig,
                    issued_to=eq_name,
                ).exists()
                if not dup:
                    HSDIssue.objects.create(
                        date=date, rig=rig,
                        purpose=purpose,
                        issued_to=eq_name,
                        equipment=eq_obj,
                        quantity_ltrs=qty,
                        meter_start=ms, meter_end=me, meter_hours=mh,
                        remarks=remarks,
                        created_by=request.user,
                    )
                    saved += 1

        if saved > 0:
            _recompute_stock(rig, date, request.user)
            messages.success(request, f'<strong>{saved}</strong> HSD issue entries saved for {rig} on {date}.')
        else:
            messages.warning(request, 'No entries saved — enter at least one quantity or meter reading.')
        return redirect('hsd_issues')

    # GET — build equipment table for selected rig
    rig_sel = request.GET.get('rig', rigs[0] if rigs else '')
    date_sel = request.GET.get('date', yesterday)

    # Get deployed equipment for this rig
    equipment_rows = []
    if rig_sel:
        # Deployed equipment
        active_deps = EquipmentDeployment.objects.filter(
            deployed_to=rig_sel, end_date__isnull=True
        ).select_related('equipment').order_by('equipment__equipment_type')

        # Define purpose mapping by equipment type
        type_purpose_map = {
            'Generator':   'Generator',
            'MudPump':     'Drilling',
            'RigCarrier':  'Drilling',
            'Pump':        'Drilling',
            'Other':       'Drilling',
            'Crane':       'Equipment',
            'Forklift':    'Equipment',
            'Hydra':       'Equipment',
            'Trailer':     'Equipment',
            'Vehicle':     'Vehicle',
            'Ambulance':   'Vehicle',
            'Camp':        'Camp',
        }

        # Get last meter readings for each equipment
        already_entered = {
            i.issued_to: i
            for i in HSDIssue.objects.filter(rig=rig_sel, date=date_sel)
        }

        for dep in active_deps:
            eq = dep.equipment
            last_issue = HSDIssue.objects.filter(
                equipment=eq, rig=rig_sel
            ).exclude(meter_end__isnull=True).order_by('-date', '-id').first()

            existing = already_entered.get(eq.equipment_no) or already_entered.get(f'{eq.equipment_no} ({eq.equipment_no})')
            equipment_rows.append({
                'eq_id':      eq.pk,
                'eq_no':      eq.equipment_no,
                'eq_type':    eq.equipment_type,
                'purpose':    type_purpose_map.get(eq.equipment_type, 'Drilling'),
                'last_meter': float(last_issue.meter_end) if last_issue and last_issue.meter_end else '',
                'existing':   existing,
            })

        # Also add ITS UNIT and other common non-equipment entries
        common_extras = [
            {'eq_id': '', 'eq_no': 'ITS UNIT',   'eq_type': 'Other', 'purpose': 'Other', 'last_meter': '', 'existing': already_entered.get('ITS UNIT')},
            {'eq_id': '', 'eq_no': 'COLD COMP',  'eq_type': 'Other', 'purpose': 'Other', 'last_meter': '', 'existing': already_entered.get('COLD COMP')},
        ]
        for ex in common_extras:
            if not any(r['eq_no'] == ex['eq_no'] for r in equipment_rows):
                equipment_rows.append(ex)

    # Get current stock
    stock = HSDDailyStock.objects.filter(rig=rig_sel, date__lte=date_sel).order_by('-date').first()

    return render(request, 'hsd/bulk_issue.html', {
        'page_title':     'Bulk HSD Issue Entry',
        'rigs':           rigs,
        'rig_sel':        rig_sel,
        'date_sel':       date_sel,
        'today':          today,
        'yesterday':      yesterday,
        'equipment_rows': equipment_rows,
        'stock':          stock,
        'purposes':       HSDIssue.PURPOSE_CHOICES,
    })

@login_required
def hsd_get_entered_equipment(request):
    """API: return equipment IDs AND issued_to names already entered for rig+date."""
    from django.http import JsonResponse
    rig  = request.GET.get('rig', '').strip()
    date = request.GET.get('date', '').strip()
    if not rig or not date:
        return JsonResponse({'entered_ids': [], 'entered_names': []})
    issues = HSDIssue.objects.filter(rig=rig, date=date).values('equipment_id', 'issued_to')
    entered_ids   = [i['equipment_id'] for i in issues if i['equipment_id']]
    entered_names = [i['issued_to'].split(' (')[0].strip() for i in issues if i['issued_to']]
    return JsonResponse({'entered_ids': entered_ids, 'entered_names': entered_names})
