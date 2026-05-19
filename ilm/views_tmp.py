import datetime
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import models
from django.db.models import Sum, Count, Q
from django.template.loader import render_to_string
from django.core.paginator import Paginator          # ? NEW
from ilm.models import ILMLog, ILMEquipmentUsage
from masters.models import Rig, WellLocation, Vendor, Equipment
from core.decorators import supervisor_required, admin_required


def _get_rigs():
    rigs = list(Rig.objects.values_list('rig_name', flat=True).order_by('rig_name'))
    return rigs if rigs else ['PPE-1', 'PPE-2', 'PPE-3', 'PPE-4', 'PPE-5']

def _get_user_rigs(request):
    all_rigs = _get_rigs()

    try:
        if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'filter_rigs'):
            return request.user.profile.filter_rigs(all_rigs)
    except Exception:
        pass

    return all_rigs
def _apply_user_rig_filter(qs, request):
    """Restrict queryset to user assigned rigs."""
    try:
        profile = request.user.profile
        if profile.role == 'admin':
            return qs
        assigned = profile.get_assigned_rigs()
        if assigned:
            return qs.filter(rig__in=assigned)
    except Exception:
        pass
    return qs


def _build_filter(request):
    p = request.POST if request.method == 'POST' else request.GET
    rig_f    = p.get('rig', '').strip()
    from_f   = p.get('from', '').strip()
    to_f     = p.get('to', datetime.date.today().isoformat()).strip()
    month_f  = p.get('month', '').strip()
    status_f = p.get('status', '').strip()

    qs = ILMLog.objects.all()
    if month_f and len(month_f) == 7:
        year, mon = month_f.split('-')
        import calendar
        last_day = calendar.monthrange(int(year), int(mon))[1]
        qs = qs.filter(date__gte=f'{month_f}-01', date__lte=f'{month_f}-{last_day:02d}')
    else:
        if from_f: qs = qs.filter(date__gte=from_f)
        if to_f:   qs = qs.filter(date__lte=to_f)
    if rig_f:    qs = qs.filter(rig=rig_f)
    if status_f: qs = qs.filter(move_status=status_f)

    return qs, {'rig': rig_f, 'from': from_f, 'to': to_f,
                'month': month_f, 'status': status_f}


@login_required
def ilm_report(request):
    qs, filters = _build_filter(request)
    qs = _apply_user_rig_filter(qs, request)

    # Aggregates run on the filtered queryset — fast, single SQL query each
    # Count unique moves (by move_group) vs raw entries
    unique_moves = qs.exclude(move_group='').values('move_group').distinct().count()
    raw_entries  = qs.filter(move_group='').count()
    total_moves  = unique_moves + raw_entries  # ungrouped count as individual

    stats = qs.aggregate(
        total_entries  = Count('id'),
        actual_moves   = Count('id', filter=Q(during_ilm_hrs__gt=0)),
        total_ilm_hrs  = Sum('during_ilm_hrs'),
        total_extra    = Sum('rig_move_extra_hrs'),
        total_saving   = Sum('rig_move_saving_hrs'),
        total_trailers = Sum('trailer_reported'),
        total_t_loss   = Sum('trailer_loss'),
    )

    rig_summary = qs.values('rig').annotate(
        entries   = Count('id'),
        total_hrs = Sum('during_ilm_hrs'),
        extra_hrs = Sum('rig_move_extra_hrs'),
        save_hrs  = Sum('rig_move_saving_hrs'),
        trailers  = Sum('trailer_reported'),
        t_loss    = Sum('trailer_loss'),
    ).order_by('rig')


    # Chart data
    import json
    status_breakdown = qs.values('move_status').annotate(
        count=Count('id'),
        hrs=Sum('during_ilm_hrs')
    ).order_by('move_status')

    status_labels = json.dumps([s['move_status'] or 'Unknown' for s in status_breakdown])
    status_counts  = json.dumps([s['count'] for s in status_breakdown])
    rig_labels     = json.dumps([r['rig'] for r in rig_summary])
    rig_hrs        = json.dumps([float(r['total_hrs'] or 0) for r in rig_summary])

    # -- PAGINATED entries — only 50 rows loaded per page ------------------
    entries_qs = qs.prefetch_related(
        'equipment_usage__equipment__vendor'
    ).order_by('-date', 'rig')

    paginator = Paginator(entries_qs, 50)
    page_num  = request.GET.get('page', 1)
    entries   = paginator.get_page(page_num)
    # ----------------------------------------------------------------------

    # Load fleet equipment for the "Add Equipment" modal
    fleet_equipment = Equipment.objects.filter(
        status__in=['Available', 'Deployed']
    ).select_related('vendor').order_by('equipment_type', 'equipment_no')

    return render(request, 'ilm/report.html', {
        'page_title':       'ILM Report',
        'entries':          entries,          # now a Page object, not full queryset
        'stats':            stats,
        'rig_summary':      rig_summary,
        'rigs':             _get_user_rigs(request),
        'filters':          filters,
        'statuses':         ['Active', 'Standby', 'Internal', 'Idle'],
        'fleet_equipment':  fleet_equipment,
        'role_choices':     ILMEquipmentUsage.ROLE_CHOICES,
        'status_labels': status_labels,
        'status_counts': status_counts,
        'rig_labels':    rig_labels,
        'rig_hrs':       rig_hrs,
    })


@login_required
@supervisor_required
def ilm_add_equipment(request, pk):
    """Add fleet equipment to an ILM log entry."""
    entry = get_object_or_404(ILMLog, pk=pk)
    if request.method == 'POST':
        eq_ids = request.POST.getlist('equipment_ids')
        role   = request.POST.get('role', 'Trailer')
        notes  = request.POST.get('notes', '').strip()
        added  = 0
        for eq_id in eq_ids:
            try:
                eq = Equipment.objects.get(pk=int(eq_id))
                ILMEquipmentUsage.objects.get_or_create(
                    ilm_log=entry, equipment=eq,
                    defaults={'role': role, 'notes': notes}
                )
                added += 1
            except Equipment.DoesNotExist:
                pass
        messages.success(request, f'{added} equipment linked to {entry.rig} / {entry.date}.')
    return redirect('ilm_report')


@login_required
@supervisor_required
def ilm_remove_equipment(request, pk, eq_pk):
    """Remove fleet equipment from an ILM log entry."""
    usage = get_object_or_404(ILMEquipmentUsage, ilm_log_id=pk, equipment_id=eq_pk)
    if request.method == 'POST':
        entry_info = f'{usage.ilm_log.rig} / {usage.ilm_log.date}'
        eq_info    = usage.equipment.equipment_no
        usage.delete()
        messages.success(request, f'{eq_info} removed from {entry_info}.')
    return redirect('ilm_report')


@login_required
@supervisor_required
def ilm_add(request):
    rigs      = _get_user_rigs(request)
    locations = list(WellLocation.objects.filter(status='Active').values(
        'location', 'category').order_by('category', 'location'))
    vendors   = list(Vendor.objects.filter(status='Active').values('vendor_code', 'vendor_name'))
    fleet_equipment = Equipment.objects.filter(
        status__in=['Available', 'Deployed']
    ).select_related('vendor').order_by('equipment_type', 'equipment_no')

    if request.method == 'POST':
        date         = request.POST.get('date', '').strip()
        rig          = request.POST.get('rig', '').strip()
        move_status  = request.POST.get('move_status', 'Active')
        from_loc     = request.POST.get('ilm_from_location', '').strip()
        to_loc       = request.POST.get('ilm_to_location', '').strip()
        dist         = request.POST.get('distance_kms', '').strip()
        exp_hrs      = request.POST.get('expected_ilm_hrs', '').strip()
        during_raw   = request.POST.get('during_ilm_hrs', '').strip()
        extra_raw    = request.POST.get('rig_move_extra_hrs', '0') or '0'
        saving_raw   = request.POST.get('rig_move_saving_hrs', '0') or '0'
        t_rep        = request.POST.get('trailer_reported', '0') or '0'
        t_loss       = request.POST.get('trailer_loss', '0') or '0'
        t_vend       = request.POST.get('trailer_vendor', '').strip()
        crane_rep    = request.POST.get('crane_reported', '').strip()
        crane_vend   = request.POST.get('crane_vendor', '').strip()
        remarks      = request.POST.get('remarks', '').strip()
        eq_ids       = request.POST.getlist('equipment_ids')
        eq_role      = request.POST.get('equipment_role', 'Trailer')

        if not date or not rig:
            messages.error(request, 'Date and rig are required.')
            return redirect('ilm_add')

        # -- Duplicate check --------------------------------------
        if ILMLog.objects.filter(date=date, rig=rig).exists():
            messages.error(request,
                f'Duplicate: An ILM entry for {rig} on {date} already exists. Use Edit instead.')
            return redirect('ilm_add')
        # ---------------------------------------------------------

        during_hrs = float(during_raw) if during_raw else None

        import datetime as _dt2
        move_group = request.POST.get('move_group', '').strip()
        if not move_group:
            try:
                date_obj2 = _dt2.date.fromisoformat(date)
                recent = ILMLog.objects.filter(
                    rig=rig,
                    ilm_from_location=from_loc,
                    ilm_to_location=to_loc,
                    move_group__gt='',
                    date__gte=date_obj2 - _dt2.timedelta(days=7),
                    date__lt=date_obj2,
                ).order_by('-date').first()
                move_group = recent.move_group if recent else f'MG-{rig}-{date}'
            except Exception:
                move_group = f'MG-{rig}-{date}'

        entry = ILMLog.objects.create(
            date=date, rig=rig, move_group=move_group, move_status=move_status,
            start_date=request.POST.get('start_date') or None,
            start_time=request.POST.get('start_time') or None,
            end_date=request.POST.get('end_date') or None,
            end_time=request.POST.get('end_time') or None,
            ilm_from_location=from_loc, ilm_to_location=to_loc,
            distance_kms=dist, expected_ilm_hrs=exp_hrs,
            during_ilm_hrs=during_hrs,
            rig_move_extra_hrs=float(extra_raw),
            rig_move_saving_hrs=float(saving_raw),
            trailer_reported=int(t_rep), trailer_loss=int(t_loss),
            trailer_vendor=t_vend, crane_reported=crane_rep,
            crane_vendor=crane_vend, remarks=remarks,
            created_by=request.user,
        )

        for eq_id in eq_ids:
            try:
                eq = Equipment.objects.get(pk=int(eq_id))
                ILMEquipmentUsage.objects.create(
                    ilm_log=entry, equipment=eq, role=eq_role)
            except Equipment.DoesNotExist:
                pass

        messages.success(request, f'ILM entry saved for <strong>{rig}</strong> on <strong>{date}</strong>.')
        return redirect('ilm_report')

    return render(request, 'ilm/add.html', {
        'page_title':       'Add ILM Entry',
        'rigs':             rigs,
        'locations':        locations,
        'vendors':          vendors,
        'fleet_equipment':  fleet_equipment,
        'role_choices':     ILMEquipmentUsage.ROLE_CHOICES,
        'today':            datetime.date.today().isoformat(),
        'default_date':      (datetime.date.today() - datetime.timedelta(days=1)).isoformat(),
        'statuses':         ['Active', 'Standby', 'Internal', 'Idle'],
    })


@login_required
@supervisor_required
def ilm_edit(request, pk):
    entry     = get_object_or_404(ILMLog, pk=pk)
    rigs      = _get_user_rigs(request)
    locations = list(WellLocation.objects.filter(status='Active').values('location', 'category'))
    vendors   = list(Vendor.objects.filter(status='Active').values('vendor_code', 'vendor_name'))
    fleet_equipment = Equipment.objects.filter(
        status__in=['Available', 'Deployed']
    ).select_related('vendor').order_by('equipment_type', 'equipment_no')
    current_equipment = entry.equipment_usage.select_related('equipment__vendor').all()

    if request.method == 'POST':
        entry.date                = request.POST.get('date', entry.date)
        entry.start_date          = request.POST.get('start_date') or None
        entry.start_time          = request.POST.get('start_time') or None
        entry.end_date            = request.POST.get('end_date') or None
        entry.end_time            = request.POST.get('end_time') or None
        entry.rig                 = request.POST.get('rig', entry.rig)
        entry.move_status         = request.POST.get('move_status', entry.move_status)
        entry.ilm_from_location   = request.POST.get('ilm_from_location', '').strip()
        entry.ilm_to_location     = request.POST.get('ilm_to_location', '').strip()
        entry.distance_kms        = request.POST.get('distance_kms', '').strip()
        entry.expected_ilm_hrs    = request.POST.get('expected_ilm_hrs', '').strip()
        during_raw                = request.POST.get('during_ilm_hrs', '').strip()
        entry.during_ilm_hrs      = float(during_raw) if during_raw else None
        entry.rig_move_extra_hrs  = float(request.POST.get('rig_move_extra_hrs', 0) or 0)
        entry.rig_move_saving_hrs = float(request.POST.get('rig_move_saving_hrs', 0) or 0)
        entry.trailer_reported    = int(request.POST.get('trailer_reported', 0) or 0)
        entry.trailer_loss        = int(request.POST.get('trailer_loss', 0) or 0)
        entry.trailer_vendor      = request.POST.get('trailer_vendor', '').strip()
        entry.crane_reported      = request.POST.get('crane_reported', '').strip()
        entry.crane_vendor        = request.POST.get('crane_vendor', '').strip()
        entry.remarks             = request.POST.get('remarks', '').strip()
        entry.save()
        messages.success(request, f'ILM entry updated for {entry.rig} on {entry.date}.')
        return redirect('ilm_report')

    return render(request, 'ilm/edit.html', {
        'page_title':          'Edit ILM Entry',
        'entry':               entry,
        'rigs':                rigs,
        'locations':           locations,
        'vendors':             vendors,
        'fleet_equipment':     fleet_equipment,
        'current_equipment':   current_equipment,
        'role_choices':        ILMEquipmentUsage.ROLE_CHOICES,
        'statuses':            ['Active', 'Standby', 'Internal', 'Idle'],
    })


@login_required
@admin_required
def ilm_delete(request, pk):
    entry = get_object_or_404(ILMLog, pk=pk)
    if request.method == 'POST':
        info = f'{entry.rig} on {entry.date}'
        entry.delete()
        messages.success(request, f'ILM entry for {info} deleted.')
    return redirect('ilm_report')


@login_required
@supervisor_required
def ilm_import(request):
    rigs    = _get_rigs()
    preview = []

    if request.method == 'POST':
        rig      = request.POST.get('rig', '').strip()
        mode     = request.POST.get('mode', 'preview')
        skip_dup = 'skip_duplicates' in request.POST
        f        = request.FILES.get('excel_file')

        if not f or not rig:
            messages.error(request, 'Please select a rig and Excel file.')
            return redirect('ilm_import')

        try:
            wb = openpyxl.load_workbook(f)
            imported = 0
            skipped  = 0
            dupes    = 0

            for ws in wb.worksheets:
                headers = [str(ws.cell(2, c).value or '').upper().strip() for c in range(1, 15)]
                has_crane_vendor = 'ILM CRANE VENDOR NAME' in headers
                remarks_col      = 14 if has_crane_vendor else 13

                for row_num in range(3, ws.max_row + 1):
                    date_val = ws.cell(row_num, 1).value
                    if not isinstance(date_val, datetime.datetime):
                        continue

                    date_obj = date_val.date()
                    if date_obj > datetime.date.today():
                        skipped += 1
                        continue

                    def cv(c):
                        v = str(ws.cell(row_num, c).value or '').strip()
                        return '' if v in ('-', 'None', '') else v

                    from_loc   = cv(2)
                    to_loc     = cv(3)
                    t_rep_raw  = cv(9)
                    during_raw = cv(6)

                    if not from_loc and not to_loc and not t_rep_raw:
                        continue

                    during_hrs = float(during_raw) if during_raw and during_raw.replace('.','').isdigit() else None
                    extra_raw  = cv(7)
                    saving_raw = cv(8)

                    if during_raw.upper().startswith('STAND'):
                        move_status = 'Standby'
                    elif from_loc.upper() == 'INTERNAL':
                        move_status = 'Internal'
                    elif from_loc or to_loc:
                        move_status = 'Active'
                    else:
                        move_status = 'Idle'

                    record = {
                        'date': date_obj, 'rig': rig,
                        'move_status': move_status,
                        'ilm_from_location': from_loc,
                        'ilm_to_location':   to_loc,
                        'distance_kms':      cv(4),
                        'expected_ilm_hrs':  cv(5),
                        'during_ilm_hrs':    during_hrs,
                        'rig_move_extra_hrs': float(extra_raw) if extra_raw and extra_raw.replace('.','').isdigit() else 0,
                        'rig_move_saving_hrs': float(saving_raw) if saving_raw and saving_raw.replace('.','').isdigit() else 0,
                        'trailer_reported': int(float(t_rep_raw)) if t_rep_raw and t_rep_raw.replace('.','').isdigit() else 0,
                        'trailer_loss':     int(float(cv(10))) if cv(10) and cv(10).replace('.','').isdigit() else 0,
                        'trailer_vendor':   cv(11),
                        'crane_reported':   cv(12),
                        'crane_vendor':     cv(13) if has_crane_vendor else '',
                        'remarks':          cv(remarks_col),
                        'sheet':            ws.title,
                    }

                    if mode == 'preview':
                        preview.append(record)
                    else:
                        exists = ILMLog.objects.filter(date=date_obj, rig=rig).exists()
                        if exists and skip_dup:
                            dupes += 1
                            continue
                        defaults = {k: v for k, v in record.items() if k != 'sheet'}
                        defaults['created_by'] = request.user
                        if exists:
                            ILMLog.objects.filter(date=date_obj, rig=rig).update(**{
                                k: v for k, v in defaults.items() if k != 'created_by'
                            })
                        else:
                            ILMLog.objects.create(**defaults)
                        imported += 1

            if mode == 'import':
                msg = f'<strong>{imported}</strong> rows imported.'
                if dupes:   msg += f' <strong>{dupes}</strong> duplicates skipped.'
                if skipped: msg += f' <strong>{skipped}</strong> future dates skipped.'
                messages.success(request, msg)
                return redirect('ilm_report')
            else:
                if preview:
                    messages.info(request, f'<strong>{len(preview)}</strong> rows found. Review below then confirm import.')
                else:
                    messages.warning(request, 'No valid rows found in this file.')

        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')

    return render(request, 'ilm/import.html', {
        'page_title': 'Import ILM Excel',
        'rigs':       rigs,
        'preview':    preview,
    })


@login_required
def ilm_export_excel(request):
    if request.method != 'POST':
        return redirect('ilm_report')

    qs, _ = _build_filter(request)
    entries = qs.prefetch_related(
        'equipment_usage__equipment'
    ).order_by('rig', 'date')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill('solid', fgColor='0B3D6D')
    hf   = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='E2E8F0')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    rigs_in_data = list(dict.fromkeys(entries.values_list('rig', flat=True)))
    for rig_name in rigs_in_data:
        ws = wb.create_sheet(title=rig_name[:31])
        headers = ['Date', 'Start Date', 'Start Time', 'End Date', 'End Time',
                   'Status', 'From', 'To', 'Dist KM', 'Exp Hrs',
                   'Actual Hrs', 'Extra Hrs', 'Saving Hrs',
                   'Trailers', 'T.Loss', 'T.Vendor',
                   'Crane', 'C.Vendor',
                   'Trailers (Reg Nos)', 'Cranes (Reg Nos)',
                   'Remarks']
        widths  = [12, 12, 12, 12, 12, 12, 18, 18, 10, 10, 10, 10, 10, 10, 8, 20, 10, 20, 30, 30, 25]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(1, col, h)
            cell.font = hf; cell.fill = NAVY
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        for idx, e in enumerate(entries.filter(rig=rig_name), 2):
            bg = PatternFill('solid', fgColor='FFFFFF' if idx % 2 == 0 else 'F8FAFC')

            eq_usage = e.equipment_usage.all()
            trailer_reg = ', '.join([
                f"{u.equipment.equipment_no}" +
                (f" [{u.equipment.registration_no}]" if u.equipment.registration_no else '')
                for u in eq_usage if u.role == 'Trailer'
            ])
            crane_reg = ', '.join([
                f"{u.equipment.equipment_no}" +
                (f" [{u.equipment.registration_no}]" if u.equipment.registration_no else '')
                for u in eq_usage if u.role == 'Crane'
            ])

            vals = [
                e.date,
                e.start_date if e.start_date else '',
                str(e.start_time)[:5] if e.start_time else '',
                e.end_date if e.end_date else '',
                str(e.end_time)[:5] if e.end_time else '',
                e.move_status,
                e.ilm_from_location, e.ilm_to_location,
                e.distance_kms, e.expected_ilm_hrs,
                float(e.during_ilm_hrs) if e.during_ilm_hrs else '',
                float(e.rig_move_extra_hrs) if e.rig_move_extra_hrs else '',
                float(e.rig_move_saving_hrs) if e.rig_move_saving_hrs else '',
                e.trailer_reported, e.trailer_loss,
                e.trailer_vendor, e.crane_reported, e.crane_vendor,
                trailer_reg, crane_reg,
                e.remarks,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(idx, col, v)
                cell.fill = bg; cell.border = bdr
                cell.alignment = Alignment(vertical='center')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ILM_Report_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def ilm_export_pdf(request):
    if request.method != 'POST':
        return redirect('ilm_report')

    # Read filters from POST (not GET)
    rig_f    = request.POST.get('rig', '').strip()
    from_f   = request.POST.get('from', '').strip()
    to_f     = request.POST.get('to', datetime.date.today().isoformat()).strip()
    month_f  = request.POST.get('month', '').strip()
    status_f = request.POST.get('status', '').strip()

    qs = ILMLog.objects.all()
    if month_f and len(month_f) == 7:
        import calendar
        year, mon = month_f.split('-')
        last_day = calendar.monthrange(int(year), int(mon))[1]
        qs = qs.filter(date__gte=f'{month_f}-01', date__lte=f'{month_f}-{last_day:02d}')
    else:
        if from_f: qs = qs.filter(date__gte=from_f)
        if to_f:   qs = qs.filter(date__lte=to_f)
    if rig_f:    qs = qs.filter(rig=rig_f)
    if status_f: qs = qs.filter(move_status=status_f)

    qs = qs.prefetch_related('equipment_usage__equipment').order_by('rig', 'date')

    filters = {'rig': rig_f, 'from': from_f, 'to': to_f,
               'month': month_f, 'status': status_f}

    stats = qs.aggregate(
        total=Count('id'), moves=Count('id', filter=models.Q(during_ilm_hrs__gt=0)),
        hrs=Sum('during_ilm_hrs'), extra=Sum('rig_move_extra_hrs'),
        saving=Sum('rig_move_saving_hrs'), trailers=Sum('trailer_reported'),
        t_loss=Sum('trailer_loss'),
    )

    unique_moves = qs.exclude(move_group='').values('move_group').distinct().count()
    raw_entries  = qs.filter(move_group='').count()
    total_moves  = unique_moves + raw_entries

    html = render_to_string('exports/ilm_pdf.html', {
        'entries':   qs,
        'stats':      stats,
        'total_moves': total_moves,
        'filters':   filters,
        'generated': datetime.datetime.now(),
    }, request=request)

    import weasyprint
    pdf = weasyprint.HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ILM_Report_{datetime.date.today()}.pdf"'
    return response