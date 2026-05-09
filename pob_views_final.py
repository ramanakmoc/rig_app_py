import datetime
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count
from django.http import JsonResponse
from core.decorators import supervisor_required, admin_required
from masters.models import Rig, WellLocation, WellLocation
from .models import (POBDailyLog, POBPerson, POBDesignation,
                     POBCompany, POBAccommodation, POBRoomNo, POBEmployee,
                     POBCategory)


def _get_rigs():
    rigs = list(Rig.objects.filter(rig_status='Active').values_list('rig_name', flat=True))
    return rigs if rigs else ['PPE-1','PPE-2','PPE-3','PPE-4','PPE-5']

def _get_user_rigs(request):
    all_rigs = _get_rigs()
    try:
        return request.user.profile.filter_rigs(all_rigs)
    except Exception:
        return all_rigs


def _get_cat_choices():
    """Safe wrapper — falls back to hardcoded list if migration not yet applied."""
    try:
        choices = POBCategory.as_choices()
        return choices if choices else POBPerson.CATEGORY_CHOICES
    except Exception:
        return POBPerson.CATEGORY_CHOICES

def _pob_qs_filter(qs, request, rig_f, from_f, to_f, month_f):
    import calendar
    if month_f and len(month_f) == 7:
        y, m = month_f.split('-')
        last = calendar.monthrange(int(y), int(m))[1]
        qs = qs.filter(date__gte=f'{month_f}-01', date__lte=f'{month_f}-{last:02d}')
    else:
        if from_f: qs = qs.filter(date__gte=from_f)
        if to_f:   qs = qs.filter(date__lte=to_f)
    if rig_f: qs = qs.filter(rig=rig_f)
    try:
        p = request.user.profile
        if p.role != 'admin':
            assigned = p.get_assigned_rigs()
            if assigned: qs = qs.filter(rig__in=assigned)
    except Exception:
        pass
    return qs


@login_required
def pob_report(request):
    rig_f   = request.GET.get('rig','').strip()
    from_f  = request.GET.get('from','').strip()
    to_f    = request.GET.get('to', datetime.date.today().isoformat()).strip()
    month_f = request.GET.get('month','').strip()

    qs = POBDailyLog.objects.prefetch_related('persons')
    qs = _pob_qs_filter(qs, request, rig_f, from_f, to_f, month_f)
    qs = qs.order_by('date','rig')

    COLORS = ['#3b82f6','#ef4444','#f59e0b','#22c55e','#8b5cf6','#14b8a6','#f97316','#06b6d4']

    # Trend chart
    rigs_in_data = list(qs.values_list('rig', flat=True).distinct().order_by('rig'))
    all_dates    = list(qs.values_list('date', flat=True).distinct().order_by('date'))
    date_labels  = json.dumps([str(d) for d in all_dates])
    trend_datasets = []
    for i, rig in enumerate(rigs_in_data):
        rig_map = {str(l.date): l.total_pob for l in qs.filter(rig=rig)}
        trend_datasets.append({
            'label': rig,
            'data':  [rig_map.get(str(d), None) for d in all_dates],
            'color': COLORS[i % len(COLORS)]
        })

    # Stats
    latest_log = qs.order_by('-date').first()
    all_logs   = list(qs)
    latest_pob = latest_log.total_pob if latest_log else 0
    latest_lti = latest_log.lti_free_days if latest_log else 0
    avg_pob    = round(sum(l.total_pob for l in all_logs) / len(all_logs), 1) if all_logs else 0
    max_pob    = max((l.total_pob for l in all_logs), default=0)

    # Rig summary for donut
    rig_summary = []
    for rig in rigs_in_data:
        rig_logs = [l for l in all_logs if l.rig == rig]
        if rig_logs:
            rig_summary.append({
                'rig':     rig,
                'avg_pob': round(sum(l.total_pob for l in rig_logs)/len(rig_logs),1),
                'max_pob': max(l.total_pob for l in rig_logs),
                'entries': len(rig_logs),
            })

    # Persons breakdown
    from pob.models import POBPerson
    log_ids    = [l.pk for l in all_logs]
    persons_qs = POBPerson.objects.filter(pob_log_id__in=log_ids, is_active=True).select_related('company')

    # Company counts for bar chart
    co_counts = {}
    for p in persons_qs:
        co = p.get_company() or 'Unknown'
        co_counts[co] = co_counts.get(co, 0) + 1
    co_counts = dict(sorted(co_counts.items(), key=lambda x: -x[1])[:8])
    cat_labels = json.dumps(list(co_counts.keys()))
    cat_values = json.dumps(list(co_counts.values()))

    # Shift data
    shift_data = json.dumps({
        'labels': ['Day','Night','General'],
        'values': [
            persons_qs.filter(shift='D').count(),
            persons_qs.filter(shift='N').count(),
            persons_qs.filter(shift='G').count(),
        ]
    })

    # Meal data
    meal_data = json.dumps({
        'labels': ['Breakfast','Lunch','Dinner'],
        'values': [
            persons_qs.filter(meal_b=True).count(),
            persons_qs.filter(meal_l=True).count(),
            persons_qs.filter(meal_d=True).count(),
        ]
    })

    return render(request, 'pob/report.html', {
        'page_title':      'POB Report',
        'logs':            qs.order_by('-date','rig'),
        'rigs':            _get_user_rigs(request),
        'filters':         {'rig':rig_f,'from':from_f,'to':to_f,'month':month_f},
        'today':           datetime.date.today().isoformat(),
        'latest_log':      latest_log,
        'latest_pob':      latest_pob,
        'latest_lti':      latest_lti,
        'avg_pob':         avg_pob,
        'max_pob':         max_pob,
        'rig_summary':     rig_summary,
        'date_labels':     date_labels,
        'trend_datasets':  json.dumps(trend_datasets),
        'cat_labels':      cat_labels,
        'cat_values':      cat_values,
        'shift_data':      shift_data,
        'meal_data':       meal_data,
        'chart_colors':    json.dumps(COLORS),
    })

def _pob_quick_groups(persons):
    """
    Returns (groups, grand_total, night_on_site) where:
      groups        = [(company_name, count), ...]  sorted by count desc
      grand_total   = total active persons
      night_on_site = night shift persons not left site
    """
    from collections import defaultdict
    company_counts = defaultdict(int)
    for p in persons:
        label = p.company.name if p.company else 'No Company'
        company_counts[label] += 1
    groups        = sorted(company_counts.items(), key=lambda x: -x[1])
    grand_total   = sum(v for _, v in groups)
    night_on_site = persons.filter(shift='N', left_site=False).count()
    return groups, grand_total, night_on_site


def pob_report_export(request):
    """PDF + Excel POB daily report."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fmt    = request.GET.get('fmt','excel')  # excel or pdf

    # Accept either ?log=<pk>  OR  ?rig=PPE-1&date=2026-05-01
    log_pk = request.GET.get('log','').strip()
    if log_pk:
        log  = get_object_or_404(POBDailyLog, pk=log_pk)
        rig  = log.rig
        date = str(log.date)
    else:
        rig  = request.GET.get('rig','').strip()
        date = request.GET.get('date','').strip()
        if not rig or not date:
            messages.error(request, 'Rig and date required.')
            return redirect('pob_report')
        try:
            log = POBDailyLog.objects.get(rig=rig, date=date)
        except POBDailyLog.DoesNotExist:
            messages.error(request, f'No POB log found for {rig} on {date}.')
            return redirect('pob_report')

    persons = log.persons.select_related('designation','company','accommodation','room_no').filter(is_active=True)
    groups, grand_total, night_on_site = _pob_quick_groups(persons)

    shift_day     = persons.filter(shift='D').count()
    shift_night   = persons.filter(shift='N').count()
    shift_general = persons.filter(shift='G').count()

    if fmt == 'pdf':
        return _pob_pdf_report(request, log, persons, groups, grand_total, night_on_site,
                               shift_day, shift_night, shift_general)

    # ── EXCEL ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    NAVY   = PatternFill('solid', fgColor='0B3D6D')
    BLUE   = PatternFill('solid', fgColor='1E5FA3')
    AMBER  = PatternFill('solid', fgColor='F59E0B')
    GREEN  = PatternFill('solid', fgColor='16A34A')
    LGRAY  = PatternFill('solid', fgColor='F1F5F9')
    LBLUE  = PatternFill('solid', fgColor='DBEAFE')

    hf_w   = Font(bold=True, color='FFFFFF', size=10)
    hf_b   = Font(bold=True, color='0B3D6D', size=10)
    bold   = Font(bold=True, size=10)
    thin   = Side(style='thin', color='CBD5E1')
    bdr    = Border(left=thin,right=thin,top=thin,bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center')

    def hdr(ws, row, col, val, fill=NAVY, font=None, align=center):
        c = ws.cell(row, col, val)
        c.fill = fill; c.font = font or hf_w
        c.alignment = align; c.border = bdr
        return c

    def cell(ws, row, col, val, fill=None, font=None, align=None):
        c = ws.cell(row, col, val)
        if fill:  c.fill  = fill
        if font:  c.font  = font
        c.alignment = align or left
        c.border = bdr
        return c

    # ── SHEET 1: QUICK VIEW ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'POB Quick View'
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 15
    ws1.row_dimensions[1].height = 35

    ws1.merge_cells('A1:B1')
    t = ws1.cell(1,1, f'TOTAL POB QUICK VIEW @ RIG {rig}  |  DATE: {date}')
    t.fill = NAVY; t.font = Font(bold=True,color='FFFFFF',size=12)
    t.alignment = center; t.border = bdr

    hdr(ws1, 2, 1, 'CATEGORY / GROUP', BLUE)
    hdr(ws1, 2, 2, 'COUNT',            BLUE)

    for i, (grp, cnt_val) in enumerate(groups, 3):
        fill = LGRAY if i % 2 == 0 else None
        cell(ws1, i, 1, grp,     fill, left_align := Alignment(horizontal='left',vertical='center'))
        c = ws1.cell(i, 2, cnt_val)
        c.alignment = center; c.border = bdr
        if fill: c.fill = fill
        if cnt_val > 0:
            c.font = Font(bold=True, color='1D4ED8')

    last = 3 + len(groups)
    ws1.merge_cells(f'A{last}:A{last}')
    cell(ws1, last,   1, 'GRAND TOTAL POB',    AMBER, Font(bold=True,color='FFFFFF',size=11), center)
    c = ws1.cell(last, 2, grand_total)
    c.fill = AMBER; c.font = Font(bold=True,color='FFFFFF',size=13)
    c.alignment = center; c.border = bdr

    cell(ws1, last+1, 1, 'TOTAL NIGHT POB LIVE (on site)', GREEN, Font(bold=True,color='FFFFFF'), center)
    c = ws1.cell(last+1, 2, night_on_site)
    c.fill = GREEN; c.font = Font(bold=True,color='FFFFFF',size=13)
    c.alignment = center; c.border = bdr

    # ── SHEET 2: SHIFT SUMMARY ───────────────────────────────────────
    ws2 = wb.create_sheet('Shift Summary')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15

    ws2.merge_cells('A1:B1')
    t2 = ws2.cell(1,1, f'TOTAL PERSONS BY SHIFT — {rig}  |  {date}')
    t2.fill = NAVY; t2.font = Font(bold=True,color='FFFFFF',size=12)
    t2.alignment = center; t2.border = bdr

    hdr(ws2, 2, 1, 'SHIFT', BLUE)
    hdr(ws2, 2, 2, 'COUNT', BLUE)

    shift_data = [
        ('DAY   (D)',     persons.filter(shift='D').count()),
        ('NIGHT (N)',     persons.filter(shift='N').count()),
        ('GENERAL (G)',   persons.filter(shift='G').count()),
        ('TOTAL',         persons.count()),
    ]
    fills = [LBLUE, PatternFill('solid',fgColor='EDE9FE'), LGRAY, AMBER]
    for i,(lbl,cnt_val) in enumerate(shift_data, 3):
        f = fills[i-3]
        cell(ws2, i, 1, lbl, f, Font(bold=True,color='1D4ED8') if lbl!='TOTAL' else Font(bold=True,color='FFFFFF'))
        c = ws2.cell(i, 2, cnt_val)
        c.fill = f; c.font = Font(bold=True, color='FFFFFF' if lbl=='TOTAL' else '1D4ED8', size=12)
        c.alignment = center; c.border = bdr

    # ── SHEET 3: FULL PERSON LIST ────────────────────────────────────
    ws3 = wb.create_sheet('Person List')
    ws3.freeze_panes = 'A3'
    col_w = [5,30,25,8,20,20,15,12,5,5,5,20]
    headers= ['#','Name','Designation','Shift','Company','Accommodation','Room','DOJ','B','L','D','Remarks']
    for i,(h,w) in enumerate(zip(headers,col_w),1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        hdr(ws3, 1, i, h, NAVY)

    ws3.merge_cells('A0:L0') if False else None
    # Title row
    ws3.insert_rows(1)
    ws3.merge_cells('A1:L1')
    tt = ws3.cell(1,1, f'POB PERSONS — {rig}  |  DATE: {date}  |  TOTAL: {grand_total}')
    tt.fill = NAVY; tt.font = Font(bold=True,color='FFFFFF',size=12)
    tt.alignment = center; tt.border = bdr
    ws3.row_dimensions[1].height = 28

    # Re-add headers on row 2
    for i,(h,w) in enumerate(zip(headers,col_w),1):
        hdr(ws3, 2, i, h, BLUE)

    for idx, p in enumerate(persons.order_by('shift','name'), 3):
        bg = PatternFill('solid', fgColor='FFFFFF' if idx%2==0 else 'F8FAFC')
        vals = [
            idx-2, p.name, p.get_designation(), p.shift,
            p.get_company(), p.get_accommodation(), p.get_room_no(),
            p.doj, '✓' if p.meal_b else '', '✓' if p.meal_l else '', '✓' if p.meal_d else '',
            p.remarks or ''
        ]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(idx, col, val)
            c.fill = bg; c.border = bdr
            c.alignment = center if col in [1,4,9,10,11] else left
            if col == 4:  # shift color
                c.font = Font(bold=True, color={
                    'D':'1D4ED8','N':'7C3AED','G':'374151'}.get(val,'374151'))

    # ── SHEET 4: NIGHT POB LIVE ──────────────────────────────────────
    ws4 = wb.create_sheet('Night POB Live')
    night_persons = persons.filter(shift='N', left_site=False).order_by('name')
    ws4.merge_cells('A1:J1')
    tt4 = ws4.cell(1,1, f'TOTAL NIGHT POB LIVE — {rig}  |  {date}  |  COUNT: {night_on_site}')
    tt4.fill = PatternFill('solid',fgColor='7C3AED')
    tt4.font = Font(bold=True,color='FFFFFF',size=12)
    tt4.alignment = center; tt4.border = bdr
    ws4.row_dimensions[1].height = 28

    n_headers = ['#','Name','Designation','Company','Accommodation','Room','DOJ','Days','Mobile','Remarks']
    n_widths  = [5,30,25,20,20,15,12,8,15,20]
    for i,(h,w) in enumerate(zip(n_headers,n_widths),1):
        ws4.column_dimensions[get_column_letter(i)].width = w
        hdr(ws4, 2, i, h, PatternFill('solid',fgColor='7C3AED'))

    for idx, p in enumerate(night_persons, 3):
        bg = PatternFill('solid', fgColor='FFFFFF' if idx%2==0 else 'F5F3FF')
        vals = [idx-2, p.name, p.get_designation(), p.get_company(),
                p.get_accommodation(), p.get_room_no(), p.doj,
                p.days_on_site, p.mobile_no or '', p.remarks or '']
        for col, val in enumerate(vals, 1):
            c = ws4.cell(idx, col, val)
            c.fill = bg; c.border = bdr
            c.alignment = center if col in [1,7,8] else left

    import io
    from django.http import HttpResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'POB_{rig}_{date}.xlsx'
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _pob_pdf_report(request, log, persons, groups, grand_total, night_on_site, shift_day=0, shift_night=0, shift_general=0):
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from django.http import HttpResponse
    html_str = render_to_string('pob/report_pdf.html', {
        'log': log, 'persons': persons, 'groups': groups,
        'grand_total': grand_total, 'night_on_site': night_on_site,
        'shift_day': shift_day, 'shift_night': shift_night, 'shift_general': shift_general,
        'company_name': 'KRISS DRILLING PVT. LTD.',
    }, request=request)
    pdf = HTML(string=html_str, base_url=request.build_absolute_uri('/')).write_pdf()
    fname = f'POB_{log.rig}_{log.date}.pdf'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def pob_day_detail(request, pk):
    log = get_object_or_404(POBDailyLog, pk=pk)
    persons = log.persons.select_related('designation','company','accommodation','room_no').order_by('shift','name')

    # ── Group by SHIFT ──────────────────────────────────────────────
    shift_groups = {'Day': [], 'Night': [], 'General': []}
    for p in persons:
        if p.shift == 'D':   shift_groups['Day'].append(p)
        elif p.shift == 'N': shift_groups['Night'].append(p)
        else:                shift_groups['General'].append(p)
    shift_groups = {k: v for k, v in shift_groups.items() if v}

    # ── Group by CATEGORY ────────────────────────────────────────────
    cat_choices_list = _get_cat_choices()                  # [(key, label), ...]
    cat_label_map    = dict(cat_choices_list)              # {key: label}
    cat_order_map    = {key: i for i, (key, _) in enumerate(cat_choices_list)}
    _cat_buckets     = {}
    for p in persons:
        key = (p.category or 'OTHER').strip()
        # Try exact match first, then case-insensitive, then use key as label
        if key in cat_label_map:
            label = cat_label_map[key]
        else:
            # Legacy / unmatched — find by case-insensitive key match
            matched = next((lbl for k, lbl in cat_choices_list if k.upper() == key.upper()), None)
            label = matched if matched else key.replace('_', ' ').title()
        if label not in _cat_buckets:
            _cat_buckets[label] = []
        _cat_buckets[label].append(p)
    # Sort by defined master order; unmatched go to end
    def _sort_key(item):
        label = item[0]
        key_for_label = next((k for k, v in cat_label_map.items() if v == label), None)
        return cat_order_map.get(key_for_label, 999)
    cat_groups = dict(sorted(_cat_buckets.items(), key=_sort_key))

    return render(request, 'pob/day_detail.html', {
        'page_title':    f'POB — {log.rig} — {log.date}',
        'log':           log,
        'categories':    shift_groups,   # legacy key — shift view
        'cat_groups':    cat_groups,     # new — category view
        'persons':       persons,
        'desigs':        POBDesignation.objects.filter(is_active=True),
        'companies':     POBCompany.objects.filter(is_active=True),
        'accomms':       POBAccommodation.objects.filter(is_active=True),
        'cat_choices':   _get_cat_choices(),
        'shift_choices': POBPerson.SHIFT_CHOICES,
    })


@login_required
@supervisor_required
def pob_add(request):
    rigs      = _get_user_rigs(request)
    today     = datetime.date.today().isoformat()
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    desigs    = list(POBDesignation.objects.filter(is_active=True).values('id','name','category'))
    companies = list(POBCompany.objects.filter(is_active=True).values('id','name'))
    accomms   = list(POBAccommodation.objects.filter(is_active=True).values('id','name'))
    rooms     = list(POBRoomNo.objects.filter(is_active=True).select_related('accommodation').values('id','room_no','accommodation__name'))

    if request.method == 'POST':
        rig      = request.POST.get('rig','').strip()
        date     = request.POST.get('date','').strip()
        location = request.POST.get('location','').strip()
        lti_days = int(request.POST.get('lti_free_days','0') or 0)
        remarks  = request.POST.get('remarks','').strip()

        if not rig or not date:
            messages.error(request, 'Rig and date required.')
            return redirect('pob_add')

        if POBDailyLog.objects.filter(rig=rig, date=date).exists():
            messages.warning(request, f'POB for {rig} on {date} already exists.')
            log = POBDailyLog.objects.get(rig=rig, date=date)
            return redirect('pob_day_detail', pk=log.pk)

        log = POBDailyLog.objects.create(
            rig=rig, date=date, location=location,
            lti_free_days=lti_days, remarks=remarks, created_by=request.user)

        _save_persons(request, log, date)
        messages.success(request, f'POB saved — {log.persons.count()} persons.')
        return redirect('pob_day_detail', pk=log.pk)

    prev_rig  = request.GET.get('rig', rigs[0] if rigs else '')
    prev_date = request.GET.get('date', yesterday)
    prev_persons = []
    try:
        prev_log = POBDailyLog.objects.filter(
            rig=prev_rig, date__lt=prev_date
        ).order_by('-date').first()
        if prev_log:
            prev_persons = list(prev_log.persons.filter(
                left_site=False, is_active=True
            ).select_related('designation','company','accommodation','room_no').order_by('shift','sno'))
    except Exception:
        pass

    return render(request, 'pob/add.html', {
        'page_title':   'Add POB Entry',
        'rigs':         rigs,
        'today':        today,
        'default_date': prev_date,
        'prev_rig':     prev_rig,
        'prev_persons': prev_persons,
        'categories':   _get_cat_choices(),
        'shifts':       POBPerson.SHIFT_CHOICES,
        'desigs':       desigs,
        'companies':    companies,
        'accomms':      accomms,
        'rooms':        rooms,
    })


def _save_persons(request, log, date):
    names      = request.POST.getlist('name')
    categories = request.POST.getlist('category')
    desig_ids  = request.POST.getlist('desig_id')
    desig_txts = request.POST.getlist('desig_text')
    shifts     = request.POST.getlist('shift')
    comp_ids   = request.POST.getlist('company_id')
    comp_txts  = request.POST.getlist('company_text')
    accomm_ids = request.POST.getlist('accommodation_id')
    accomm_txts= request.POST.getlist('accommodation_text')
    room_ids   = request.POST.getlist('room_id')
    room_txts  = request.POST.getlist('room_text')
    doj_list   = request.POST.getlist('doj')
    mobiles    = request.POST.getlist('mobile_no')
    meal_b_set = set(request.POST.getlist('meal_b'))
    meal_l_set = set(request.POST.getlist('meal_l'))
    meal_d_set = set(request.POST.getlist('meal_d'))
    arrived_set= set(request.POST.getlist('arrived'))
    left_set   = set(request.POST.getlist('left_site'))
    rem_list   = request.POST.getlist('person_remarks')
    date_obj   = datetime.date.fromisoformat(date)

    for i, name in enumerate(names):
        name = name.strip()
        if not name: continue
        doj_val = None
        if i < len(doj_list) and doj_list[i].strip():
            try: doj_val = datetime.date.fromisoformat(doj_list[i].strip())
            except ValueError: pass
        days = (date_obj - doj_val).days + 1 if doj_val else 0
        desig_obj = None
        did = desig_ids[i].strip() if i < len(desig_ids) else ''
        if did:
            try: desig_obj = POBDesignation.objects.get(pk=int(did))
            except: pass
        comp_obj = None
        cid = comp_ids[i].strip() if i < len(comp_ids) else ''
        if cid:
            try: comp_obj = POBCompany.objects.get(pk=int(cid))
            except: pass
        accomm_obj = None
        aid = accomm_ids[i].strip() if i < len(accomm_ids) else ''
        if aid:
            try: accomm_obj = POBAccommodation.objects.get(pk=int(aid))
            except: pass
        room_obj = None
        rid = room_ids[i].strip() if i < len(room_ids) else ''
        if rid:
            try: room_obj = POBRoomNo.objects.get(pk=int(rid))
            except: pass
        POBPerson.objects.create(
            pob_log=log, sno=i+1,
            category=categories[i] if i < len(categories) else 'KSD_CREW',
            name=name, designation=desig_obj,
            designation_text=desig_txts[i] if i < len(desig_txts) else '',
            shift=shifts[i] if i < len(shifts) else 'G',
            company=comp_obj,
            company_text=comp_txts[i] if i < len(comp_txts) else '',
            accommodation=accomm_obj,
            accommodation_text=accomm_txts[i] if i < len(accomm_txts) else '',
            room_no=room_obj,
            room_no_text=room_txts[i] if i < len(room_txts) else '',
            doj=doj_val, days_on_site=days,
            mobile_no=mobiles[i] if i < len(mobiles) else '',
            meal_b=str(i) in meal_b_set, meal_l=str(i) in meal_l_set,
            meal_d=str(i) in meal_d_set, arrived=str(i) in arrived_set,
            left_site=str(i) in left_set,
            remarks=rem_list[i] if i < len(rem_list) else '',
        )


@login_required
@supervisor_required
def pob_add_person(request, log_pk):
    log = get_object_or_404(POBDailyLog, pk=log_pk)
    if request.method == 'POST':
        name = request.POST.get('name','').strip()
        if name:
            _save_persons(request, log, str(log.date))
            messages.success(request, 'Person added to POB.')
    return redirect('pob_day_detail', pk=log_pk)


@login_required
@supervisor_required
def pob_edit_person(request, pk):
    p = get_object_or_404(POBPerson, pk=pk)
    if request.method == 'POST':
        p.name      = request.POST.get('name', p.name).strip()
        p.category  = request.POST.get('category', p.category)
        p.shift     = request.POST.get('shift', 'G')
        p.mobile_no = request.POST.get('mobile_no','').strip()
        p.remarks   = request.POST.get('remarks','').strip()
        p.meal_b    = 'meal_b'    in request.POST
        p.meal_l    = 'meal_l'    in request.POST
        p.meal_d    = 'meal_d'    in request.POST
        p.arrived   = 'arrived'   in request.POST
        p.left_site = 'left_site' in request.POST
        p.is_active = 'is_active' in request.POST
        did = request.POST.get('desig_id','').strip()
        p.designation = POBDesignation.objects.get(pk=int(did)) if did else None
        p.designation_text = request.POST.get('desig_text','').strip()
        cid = request.POST.get('company_id','').strip()
        p.company = POBCompany.objects.get(pk=int(cid)) if cid else None
        p.company_text = request.POST.get('company_text','').strip()
        aid = request.POST.get('accommodation_id','').strip()
        p.accommodation = POBAccommodation.objects.get(pk=int(aid)) if aid else None
        p.accommodation_text = request.POST.get('accommodation_text','').strip()
        rid = request.POST.get('room_id','').strip()
        p.room_no = POBRoomNo.objects.get(pk=int(rid)) if rid else None
        p.room_no_text = request.POST.get('room_text','').strip()
        doj_raw = request.POST.get('doj','').strip()
        if doj_raw:
            try:
                p.doj = datetime.date.fromisoformat(doj_raw)
                p.days_on_site = (p.pob_log.date - p.doj).days + 1
            except ValueError: pass
        p.save()
        messages.success(request, f'{p.name} updated.')
    return redirect('pob_day_detail', pk=p.pob_log.pk)


@login_required
@admin_required
def pob_delete_person(request, pk):
    p = get_object_or_404(POBPerson, pk=pk)
    log_pk = p.pob_log.pk
    if request.method == 'POST':
        name = p.name; p.delete()
        messages.success(request, f'{name} removed.')
    return redirect('pob_day_detail', pk=log_pk)


@login_required
@admin_required
def pob_delete_log(request, pk):
    log = get_object_or_404(POBDailyLog, pk=pk)
    if request.method == 'POST':
        info = f'{log.rig} on {log.date}'; log.delete()
        messages.success(request, f'POB log for {info} deleted.')
    return redirect('pob_report')


@login_required
@supervisor_required
def pob_masters(request):
    try:
        if not POBCategory.objects.exists():
            POBCategory.seed_defaults()
        pob_categories = POBCategory.objects.all()
    except Exception:
        pob_categories = []
    return render(request, 'pob/masters.html', {
        'page_title':      'POB Masters',
        'desigs':          POBDesignation.objects.all(),
        'companies':       POBCompany.objects.all(),
        'accomms':         POBAccommodation.objects.prefetch_related('rooms').all(),
        'pob_categories':  pob_categories,
    })


@login_required
@supervisor_required
def pob_master_save(request, master_type):
    if request.method != 'POST':
        return redirect('pob_masters')
    if master_type == 'designation':
        pk=request.POST.get('pk','').strip(); name=request.POST.get('name','').strip(); cat=request.POST.get('category','').strip()
        if pk:
            obj=get_object_or_404(POBDesignation,pk=pk); obj.name=name; obj.category=cat; obj.save()
        else:
            POBDesignation.objects.get_or_create(name=name, defaults={'category':cat})
    elif master_type == 'company':
        pk=request.POST.get('pk','').strip(); name=request.POST.get('name','').strip(); code=request.POST.get('short_code','').strip()
        if pk:
            obj=get_object_or_404(POBCompany,pk=pk); obj.name=name; obj.short_code=code; obj.save()
        else:
            POBCompany.objects.get_or_create(name=name, defaults={'short_code':code})
    elif master_type == 'accommodation':
        pk=request.POST.get('pk','').strip(); name=request.POST.get('name','').strip()
        if pk:
            obj=get_object_or_404(POBAccommodation,pk=pk); obj.name=name; obj.save()
        else:
            POBAccommodation.objects.get_or_create(name=name)
    elif master_type == 'room':
        pk=request.POST.get('pk','').strip(); accomm_id=request.POST.get('accommodation_id','').strip(); room_no=request.POST.get('room_no','').strip()
        accomm=get_object_or_404(POBAccommodation,pk=accomm_id) if accomm_id else None
        if pk:
            obj=get_object_or_404(POBRoomNo,pk=pk); obj.room_no=room_no; obj.accommodation=accomm; obj.save()
        else:
            POBRoomNo.objects.get_or_create(accommodation=accomm,room_no=room_no)
    elif master_type == 'category':
        pk    = request.POST.get('pk','').strip()
        label = request.POST.get('label','').strip()
        key   = request.POST.get('key','').strip().upper().replace(' ','_')
        order = int(request.POST.get('sort_order', 0) or 0)
        active = request.POST.get('is_active','') == 'on'
        if not label:
            messages.error(request, 'Label is required.')
            return redirect('pob_masters')
        if pk:
            obj = get_object_or_404(POBCategory, pk=pk)
            obj.label = label; obj.sort_order = order; obj.is_active = active; obj.save()
        else:
            if not key:
                import re
                key = re.sub(r'[^A-Z0-9]+', '_', label.upper()).strip('_')[:50]
            if POBCategory.objects.filter(key=key).exists():
                messages.error(request, f'A category with key "{key}" already exists.')
                return redirect('pob_masters')
            POBCategory.objects.create(key=key, label=label, sort_order=order, is_active=True)
    messages.success(request, 'Saved.')
    return redirect('pob_masters')


@login_required
@admin_required
def pob_master_delete(request, master_type, pk):
    if request.method == 'POST':
        model_map = {'designation':POBDesignation,'company':POBCompany,'accommodation':POBAccommodation,'room':POBRoomNo,'category':POBCategory}
        model = model_map.get(master_type)
        if model:
            obj=get_object_or_404(model,pk=pk); name=str(obj); obj.delete()
            messages.success(request, f'"{name}" deleted.')
    return redirect('pob_masters')


@login_required
def pob_api_rooms(request):
    accomm_id = request.GET.get('accommodation_id','')
    rooms = POBRoomNo.objects.filter(is_active=True)
    if accomm_id: rooms = rooms.filter(accommodation_id=accomm_id)
    return JsonResponse({'rooms': [{'id':r.pk,'room_no':r.room_no} for r in rooms]})


@login_required
def pob_employees(request):
    rig_f   = request.GET.get('rig','').strip()
    comp_f  = request.GET.get('company','').strip()
    desig_f = request.GET.get('desig','').strip()
    shift_f = request.GET.get('shift','').strip()
    cat_f   = request.GET.get('category','').strip()
    qs = POBEmployee.objects.select_related('designation','company').filter(is_active=True)
    if rig_f:   qs = qs.filter(rig=rig_f)
    if comp_f:  qs = qs.filter(company_id=comp_f)
    if desig_f: qs = qs.filter(designation_id=desig_f)
    if shift_f: qs = qs.filter(shift=shift_f)
    if cat_f:   qs = qs.filter(category=cat_f)
    cat_choices = _get_cat_choices()
    cat_dict    = dict(cat_choices)
    return render(request, 'pob/employees.html', {
        'page_title':  'POB Employee Master',
        'employees':   qs.order_by('rig','name'),
        'companies':   POBCompany.objects.filter(is_active=True),
        'desigs':      POBDesignation.objects.filter(is_active=True),
        'rigs':        _get_rigs(),
        'filters':     {'rig':rig_f,'company':comp_f,'desig':desig_f,'shift':shift_f,'category':cat_f},
        'cat_choices': cat_choices,
        'cat_dict':    cat_dict,
    })


@login_required
@supervisor_required
def pob_employee_save(request):
    if request.method != 'POST':
        return redirect('pob_employees')
    try:
        pk       = request.POST.get('pk','').strip()
        name     = request.POST.get('name','').strip()
        rig      = request.POST.get('rig','').strip()
        desig_id = request.POST.get('desig_id','').strip()
        comp_id  = request.POST.get('company_id','').strip()
        shift    = request.POST.get('shift','G')
        category = request.POST.get('category','KSD_CREW')
        mobile   = request.POST.get('mobile_no','').strip()

        if not name:
            messages.error(request, 'Name is required.')
            return redirect('pob_employees')

        desig = POBDesignation.objects.get(pk=int(desig_id)) if desig_id else None
        comp  = POBCompany.objects.get(pk=int(comp_id))      if comp_id  else None

        if pk:
            obj = get_object_or_404(POBEmployee, pk=pk)
            obj.name=name; obj.rig=rig; obj.designation=desig; obj.company=comp
            obj.shift=shift; obj.category=category; obj.mobile_no=mobile
            obj.save()
            messages.success(request, f'{name} updated.')
        else:
            # Check for duplicate (name + company unique_together)
            if POBEmployee.objects.filter(name=name, company=comp).exists():
                messages.error(request, f'Employee "{name}" already exists for this company.')
                return redirect('pob_employees')
            POBEmployee.objects.create(
                name=name, rig=rig, designation=desig, company=comp,
                shift=shift, category=category, mobile_no=mobile
            )
            messages.success(request, f'{name} added.')
    except Exception as e:
        messages.error(request, f'Error saving employee: {e}')
    return redirect('pob_employees')


@login_required
@admin_required
def pob_employee_delete(request, pk):
    emp = get_object_or_404(POBEmployee, pk=pk)
    if request.method == 'POST':
        name = emp.name; emp.delete()
        messages.success(request, f'{name} deleted.')
    return redirect('pob_employees')


@login_required
def pob_api_employees(request):
    q        = request.GET.get('q','').strip()
    rig      = request.GET.get('rig','').strip()
    category = request.GET.get('category','').strip()
    date     = request.GET.get('date','').strip()
    qs = POBEmployee.objects.filter(is_active=True).select_related('designation','company')
    if q:        qs = qs.filter(name__icontains=q)
    if rig:      qs = qs.filter(rig=rig)
    if category: qs = qs.filter(category=category)
    if rig and date:
        # Exclude by (name, company) pair so same-named people from different companies can still be added
        already_pairs = set(
            POBPerson.objects.filter(pob_log__rig=rig, pob_log__date=date)
            .values_list('name', 'company_id')
        )
        if already_pairs:
            from django.db.models import Q
            excl = Q()
            for _name, _comp_id in already_pairs:
                if _comp_id:
                    excl |= Q(name=_name, company_id=_comp_id)
                else:
                    excl |= Q(name=_name, company__isnull=True)
            qs = qs.exclude(excl)
    qs = qs.order_by('name')[:50]
    data = [{'id':e.pk,'name':e.name,
             'desig_id':e.designation.pk if e.designation else '',
             'desig':e.designation.name if e.designation else '',
             'comp_id':e.company.pk if e.company else '',
             'company':e.company.name if e.company else '',
             'shift':e.shift,'mobile':e.mobile_no,'category':e.category}
            for e in qs]
    return JsonResponse({'employees': data})
