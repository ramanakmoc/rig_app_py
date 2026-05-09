"""
Management command: import_employees_excel
Usage:
    python manage.py import_employees_excel <xlsx_path> <rig_name>

Example:
    python manage.py import_employees_excel /tmp/POB_PPE-2-2026.xlsx PPE-2
    python manage.py import_employees_excel /tmp/POB_PPE-3-2026.xlsx PPE-3
"""

import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


# ── Section header → category key map ─────────────────────────────────────────
CATEGORY_MAP = {
    'VEDANTA PERSON':                'VEDANTA_PERSON',
    'VEDANTA PERSONS':               'VEDANTA_PERSON',
    'VEDANTA VISITORS':              'VEDANTA_VISITOR',
    'VEDANTA VISITOR':               'VEDANTA_VISITOR',
    'VEDANTA SERVICES':              'VEDANTA_SERVICE',
    'VEDANTA SERVICE':               'VEDANTA_SERVICE',
    'VCIL SERVICES':                 'VEDANTA_SERVICE',
    'VEDANTA SERVICE PARTY DRIVE':   'VEDANTA_DRIVER',
    'VEDANTA SERVICE PARTY DRIVER':  'VEDANTA_DRIVER',
    'VEDANTA DRIVER':                'VEDANTA_DRIVER',
    'GENERAL SHIFT CREW':            'KSD_CREW',
    'DAY SHIFT CREW':                'KSD_CREW',
    'NIGHT SHIFT CREW':              'NIGHT_SHIFT_CREW',
    'KSD 3RD PARTY':                 'KSD_3RD_PARTY',
    'KSD THIRD PARTY':               'KSD_3RD_PARTY',
    'KSD 3RD PARTY CREW':            'KSD_3RD_PARTY',
    'KSD VISITORS':                  'KSD_VISITORS',
    "KSD'S VISITORS":                'KSD_VISITORS',
    "KSD'S VISITOR":                 'KSD_VISITORS',
    "KSD'S DRIVER":                  'KSD_VISITORS',
    "KSD'S DRIVERS":                 'KSD_VISITORS',
    'CATERING SERVICES AT RIG SITE': 'CATERING_SERVICES_AT_RIG_SITE',
    'CATERING SERVICES AT RIG':      'CATERING_SERVICES_AT_RIG_SITE',
    'CATERING STAFF AT RIG':         'CATERING_SERVICES_AT_RIG_SITE',
    'ILM CRANE':                     'KSD_3RD_PARTY',
    'ILM TRAILER':                   'KSD_3RD_PARTY',
    'ILM TLR DVR & HELP':            'KSD_3RD_PARTY',
    'ILM TLR DVR & HELPS':           'KSD_3RD_PARTY',
    'SECURITY':                      'SECURITY',
}

SHIFT_MAP = {'D': 'D', 'N': 'N', 'G': 'G', 'DAY': 'D', 'NIGHT': 'N', 'GENERAL': 'G'}

# Columns: 0=SNo, 1=Name, 2=Designation, 3=Shift, 4=Company, 5=Accommodation, 6=Room, 7=DOJ, 8=Days, 9=Nationality, 10=Mobile
COL_NAME    = 1
COL_DESIG   = 2
COL_SHIFT   = 3
COL_COMPANY = 4
COL_MOBILE  = 10


def _clean(val):
    """Strip string, return empty string for None/formula."""
    if val is None:
        return ''
    s = str(val).strip()
    if s.startswith('='):
        return ''
    return s


def _is_empty_sno(sno):
    """Treat None AND whitespace-only strings as blank serial number."""
    if sno is None: return True
    if isinstance(sno, str) and sno.strip() == '': return True
    return False


def _is_section_header(row):
    """Row where col 0 is empty/None and col 1 is a text label (category name)."""
    sno  = row[0]
    name = _clean(row[1]) if len(row) > 1 else ''
    if not _is_empty_sno(sno):
        return False, None
    if not name or name.startswith('='):
        return False, None
    key = CATEGORY_MAP.get(name.upper().strip())
    if key:
        return True, key
    # Fuzzy: check if any known key is contained in name
    for known, cat_key in CATEGORY_MAP.items():
        if known in name.upper().strip():
            return True, cat_key
    return False, None


def _is_data_row(row):
    """Row where col 0 is an integer (serial number) and col 1 has a name."""
    sno  = row[0]
    name = _clean(row[1]) if len(row) > 1 else ''
    return isinstance(sno, (int, float)) and bool(name) and not name.startswith('=')


def parse_excel(path):
    """Parse Excel file and return list of employee dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise CommandError('openpyxl not installed. Run: pip install openpyxl --break-system-packages')

    wb = load_workbook(path, read_only=True, data_only=True)
    # Use the first sheet only (first date = base employee list)
    ws = wb[wb.sheetnames[0]]

    employees = []
    current_category = 'KSD_CREW'  # default

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue

        is_hdr, cat_key = _is_section_header(row)
        if is_hdr:
            current_category = cat_key
            continue

        if _is_data_row(row):
            name  = _clean(row[COL_NAME]).upper()
            desig = _clean(row[COL_DESIG]).upper() if len(row) > COL_DESIG else ''
            shift = SHIFT_MAP.get((_clean(row[COL_SHIFT]) if len(row) > COL_SHIFT else '').upper(), 'G')
            comp  = _clean(row[COL_COMPANY]).upper() if len(row) > COL_COMPANY else ''
            mob   = _clean(row[COL_MOBILE]) if len(row) > COL_MOBILE else ''
            # Skip summary rows or formula-only rows
            if not name or name.startswith('='):
                continue
            employees.append({
                'name':        name,
                'designation': desig,
                'shift':       shift,
                'company':     comp,
                'mobile':      mob,
                'category':    current_category,
            })

    return employees


class Command(BaseCommand):
    help = 'Import employees from Excel POB sheet into POB Employee Master'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', help='Path to the Excel file')
        parser.add_argument('rig',       help='Rig name, e.g. PPE-2')
        parser.add_argument('--dry-run', action='store_true',
                            help='Preview what would be imported without saving')

    def handle(self, *args, **options):
        from pob.models import POBEmployee, POBDesignation, POBCompany

        path    = options['xlsx_path']
        rig     = options['rig'].strip().upper()
        dry_run = options['dry_run']

        self.stdout.write(f'\n📂  Reading: {path}')
        employees = parse_excel(path)
        self.stdout.write(f'✅  Found {len(employees)} employees in Excel\n')

        created = updated = skipped = 0

        for emp in employees:
            name       = emp['name']
            desig_name = emp['designation']
            comp_name  = emp['company']

            if not name:
                continue

            # Get or create Designation
            desig = None
            if desig_name:
                desig, _ = POBDesignation.objects.get_or_create(
                    name__iexact=desig_name,
                    defaults={'name': desig_name, 'is_active': True}
                )

            # Get or create Company
            comp = None
            if comp_name:
                comp, _ = POBCompany.objects.get_or_create(
                    name__iexact=comp_name,
                    defaults={'name': comp_name, 'is_active': True}
                )

            # Check by name + company only (unique_together constraint)
            existing = POBEmployee.objects.filter(
                name__iexact=name,
                company=comp,
            ).first()

            if dry_run:
                action = 'UPDATE' if existing else 'CREATE'
                self.stdout.write(
                    f'  [{action}] {name:<30} | {desig_name:<25} | '
                    f'{comp_name:<20} | {emp["shift"]} | {emp["category"]}'
                )
                if existing:
                    updated += 1
                else:
                    created += 1
                continue

            if existing:
                existing.designation = desig
                existing.shift       = emp['shift']
                existing.category    = emp['category']
                existing.rig         = rig
                existing.is_active   = True
                if emp['mobile']:
                    existing.mobile_no = emp['mobile']
                existing.save()
                updated += 1
                self.stdout.write(f"  ✏️   Updated : {name} (already in master)")
            else:
                try:
                    POBEmployee.objects.create(
                        name        = name,
                        rig         = rig,
                        designation = desig,
                        company     = comp,
                        shift       = emp['shift'],
                        category    = emp['category'],
                        mobile_no   = emp['mobile'] or '',
                        is_active   = True,
                    )
                    created += 1
                    self.stdout.write(f'  ✅  Created : {name}')
                except Exception as e:
                    skipped += 1
                    self.stdout.write(f'  ⚠️   Skipped : {name} — {e}')

        self.stdout.write('\n' + '─' * 60)
        if dry_run:
            self.stdout.write(f'🔍  DRY RUN — nothing saved')
            self.stdout.write(f'    Would create : {created}')
            self.stdout.write(f'    Would update : {updated}')
        else:
            self.stdout.write(f'🎉  Import complete for RIG {rig}')
            self.stdout.write(f'    Created : {created}')
            self.stdout.write(f'    Updated : {updated}')
            self.stdout.write(f'    Skipped : {skipped}')
        self.stdout.write('─' * 60 + '\n')
