import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from core.models import RigDailyLog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _navy():  return PatternFill('solid', fgColor='0B3D6D')
def _hdr_font(): return Font(bold=True, color='FFFFFF', size=10)
def _thin_border():
    s = Side(style='thin', color='E2E8F0')
    return Border(left=s, right=s, top=s, bottom=s)


def _build_qs(request):
    rig_f  = request.POST.get('rig', '').strip()
    from_f = request.POST.get('from', '').strip()
    to_f   = request.POST.get('to', datetime.date.today().isoformat()).strip()
    qs = RigDailyLog.objects.all()
    if rig_f:  qs = qs.filter(rig=rig_f)
    if from_f: qs = qs.filter(date__gte=from_f)
    if to_f:   qs = qs.filter(date__lte=to_f)
    return qs.order_by('date', 'rig')


@login_required
def export_excel(request):
    if request.method != 'POST':
        from django.shortcuts import redirect
        return redirect('daily_report')

    qs = _build_qs(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Log'

    headers = ['Date', 'Rig', 'Status', 'Operating Hrs', 'Standby Hrs',
               'Breakdown Hrs', 'ILM Hrs', 'Zero Rate Hrs', 'Total Hrs',
               'Efficiency %', 'Reason', 'Created By']
    widths  = [12, 10, 12, 14, 12, 14, 10, 14, 10, 12, 30, 15]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col, h)
        cell.font   = _hdr_font()
        cell.fill   = _navy()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    for row_idx, e in enumerate(qs, 2):
        bg = 'FFFFFF' if row_idx % 2 == 0 else 'F8FAFC'
        fill = PatternFill('solid', fgColor=bg)
        total = float(e.operating_hours) + float(e.standby_hours) + \
                float(e.breakdown_hours) + float(e.ilm_hours) + float(e.zero_rate_hours)
        eff   = round(float(e.operating_hours) / total * 100, 1) if total > 0 else 0
        vals  = [
            e.date, e.rig, e.status,
            float(e.operating_hours), float(e.standby_hours),
            float(e.breakdown_hours), float(e.ilm_hours), float(e.zero_rate_hours),
            round(total, 2), eff,
            e.reason, e.created_by.username if e.created_by else '',
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row_idx, col, v)
            cell.fill   = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical='center')
            if col == 1 and isinstance(v, datetime.date):
                cell.number_format = 'DD-MMM-YYYY'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rig_log_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf_daily(request):
    if request.method != 'POST':
        from django.shortcuts import redirect
        return redirect('daily_report')

    qs   = _build_qs(request)
    html = render_to_string('exports/daily_pdf.html', {
        'entries':    qs,
        'generated':  datetime.datetime.now(),
        'filters':    request.POST,
    }, request=request)

    import weasyprint
    pdf = weasyprint.HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="daily_report_{datetime.date.today()}.pdf"'
    return response


@login_required
def export_pdf_weekly(request):
    from django.shortcuts import redirect
    if request.method != 'POST':
        return redirect('weekly_report')
    from django.db.models.functions import TruncWeek
    qs = _build_qs(request)
    weekly = qs.annotate(week=TruncWeek('date')).values('week', 'rig').annotate(
        op_hrs=Sum('operating_hours'), sb_hrs=Sum('standby_hours'),
        bd_hrs=Sum('breakdown_hours'), ilm_hrs=Sum('ilm_hours'),
        zr_hrs=Sum('zero_rate_hours'), entries=Count('id'),
    ).order_by('-week', 'rig')

    html = render_to_string('exports/weekly_pdf.html', {
        'weekly': weekly, 'generated': datetime.datetime.now()
    }, request=request)
    import weasyprint
    pdf = weasyprint.HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="weekly_report_{datetime.date.today()}.pdf"'
    return response


@login_required
def export_pdf_monthly(request):
    from django.shortcuts import redirect
    if request.method != 'POST':
        return redirect('monthly_report')
    from django.db.models.functions import TruncMonth
    qs = _build_qs(request)
    monthly = qs.annotate(month=TruncMonth('date')).values('month', 'rig').annotate(
        op_hrs=Sum('operating_hours'), sb_hrs=Sum('standby_hours'),
        bd_hrs=Sum('breakdown_hours'), ilm_hrs=Sum('ilm_hours'),
        zr_hrs=Sum('zero_rate_hours'), entries=Count('id'),
    ).order_by('-month', 'rig')

    html = render_to_string('exports/monthly_pdf.html', {
        'monthly': monthly, 'generated': datetime.datetime.now()
    }, request=request)
    import weasyprint
    pdf = weasyprint.HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="monthly_report_{datetime.date.today()}.pdf"'
    return response
