import csv
import hashlib
import io
import json
import zipfile
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone

from email_reports.models import EmailAttachment


TEXT_EXTENSIONS = {".txt", ".log", ".md"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}


def _read_bytes(attachment):
    with attachment.file.open("rb") as handle:
        return handle.read()


def _tabular_text(rows, max_rows=10000):
    output = []
    for index, row in enumerate(rows):
        if index >= max_rows:
            break
        output.append(" | ".join("" if value is None else str(value) for value in row))
    return "\n".join(output)


def _extract_csv(data):
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return _tabular_text(rows), {"rows": len(rows), "format": "csv"}


def _extract_xlsx(data):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = {}
    text_parts = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        sheets[worksheet.title] = {"rows": len(rows), "columns": max((len(row) for row in rows), default=0)}
        text_parts.append(f"[Sheet: {worksheet.title}]\n{_tabular_text(rows)}")
    workbook.close()
    return "\n\n".join(text_parts), {"format": "xlsx", "sheets": sheets}


def _extract_xls(data):
    import xlrd

    workbook = xlrd.open_workbook(file_contents=data)
    sheets = {}
    text_parts = []
    for worksheet in workbook.sheets():
        rows = [worksheet.row_values(row_index) for row_index in range(worksheet.nrows)]
        sheets[worksheet.name] = {"rows": worksheet.nrows, "columns": worksheet.ncols}
        text_parts.append(f"[Sheet: {worksheet.name}]\n{_tabular_text(rows)}")
    return "\n\n".join(text_parts), {"format": "xls", "sheets": sheets}


def _ocr_image(image):
    import pytesseract

    return pytesseract.image_to_string(image)


def _extract_image(data):
    from PIL import Image

    with Image.open(io.BytesIO(data)) as image:
        text = _ocr_image(image)
        metadata = {"format": image.format, "width": image.width, "height": image.height, "ocr": True}
    return text, metadata


def _extract_pdf(data):
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    metadata = {"format": "pdf", "pages": len(reader.pages), "ocr": False}
    minimum_text = int(getattr(settings, "EMAIL_COLLECTION_PDF_OCR_MIN_CHARS", 40))
    if len(text.strip()) < minimum_text:
        from pdf2image import convert_from_bytes

        images = convert_from_bytes(data, dpi=200)
        text = "\n".join(_ocr_image(image) for image in images)
        metadata["ocr"] = True
    return text, metadata


def _safe_member_name(name):
    path = Path(name)
    if path.is_absolute() or ".." in path.parts:
        return ""
    return Path(path.name).name[:240]


def _extract_zip(attachment, data):
    max_members = int(getattr(settings, "EMAIL_COLLECTION_MAX_ZIP_MEMBERS", 50))
    max_bytes = int(getattr(settings, "EMAIL_COLLECTION_MAX_ZIP_UNCOMPRESSED_BYTES", 100 * 1024 * 1024))
    created = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        members = [member for member in archive.infolist() if not member.is_dir()]
        if len(members) > max_members:
            raise ValueError(f"ZIP contains more than {max_members} files.")
        if sum(member.file_size for member in members) > max_bytes:
            raise ValueError("ZIP uncompressed size exceeds the configured limit.")
        for member in members:
            filename = _safe_member_name(member.filename)
            if not filename:
                continue
            payload = archive.read(member)
            child = EmailAttachment(
                message=attachment.message,
                parent_archive=attachment,
                filename=filename,
                content_type="application/octet-stream",
                size_bytes=len(payload),
                checksum_sha256=hashlib.sha256(payload).hexdigest(),
            )
            child.file.save(filename, ContentFile(payload), save=False)
            child.save()
            process_attachment(child)
            created.append({"id": child.pk, "filename": filename, "status": child.status})
    return "\n".join(
        child.extracted_text for child in attachment.archive_members.all() if child.extracted_text
    ), {"format": "zip", "members": created}


def process_attachment(attachment):
    if attachment.status == "invalid":
        return attachment
    extension = Path(attachment.filename).suffix.lower()
    data = _read_bytes(attachment)
    try:
        if extension == ".csv":
            text, metadata = _extract_csv(data)
        elif extension == ".xlsx":
            text, metadata = _extract_xlsx(data)
        elif extension == ".xls":
            text, metadata = _extract_xls(data)
        elif extension == ".pdf":
            text, metadata = _extract_pdf(data)
        elif extension == ".zip":
            text, metadata = _extract_zip(attachment, data)
        elif extension in IMAGE_EXTENSIONS:
            text, metadata = _extract_image(data)
        elif extension in TEXT_EXTENSIONS:
            text = data.decode("utf-8-sig", errors="replace")
            metadata = {"format": extension.lstrip(".")}
        else:
            attachment.status = "unsupported"
            attachment.error_message = f"Unsupported attachment type: {extension or 'unknown'}"
            attachment.processed_at = timezone.now()
            attachment.save(update_fields=["status", "error_message", "processed_at"])
            return attachment
        attachment.extracted_text = text[:2_000_000]
        attachment.extracted_data = metadata
        attachment.status = "extracted"
        attachment.error_message = ""
    except Exception as exc:
        attachment.status = "failed"
        attachment.error_message = str(exc)[:2000]
    attachment.processed_at = timezone.now()
    attachment.save(
        update_fields=["extracted_text", "extracted_data", "status", "error_message", "processed_at"]
    )
    return attachment

